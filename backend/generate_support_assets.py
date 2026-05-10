from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook


def make_pdf(path: Path, title: str, audience: str, sections: list[tuple[str, list[str]]]):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    pdf.set_fill_color(24, 58, 120)
    pdf.rect(0, 0, 210, 30, style="F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_xy(12, 10)
    pdf.cell(0, 8, title)

    pdf.set_text_color(30, 30, 30)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_xy(12, 34)
    pdf.multi_cell(0, 7, f"Program documentation for {audience}. IIT-G Parent Connect workflow guide.")

    for section_title, bullets in sections:
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(18, 54, 99)
        pdf.cell(0, 8, section_title)
        pdf.ln(8)
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(30, 30, 30)
        for bullet in bullets:
            pdf.multi_cell(0, 7, f"- {bullet}")
        pdf.ln(1)

    path.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(path))


def make_student_template(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["RNO", "NAME", "WNO"])
    ws.append(["23AIDS001", "Student One", "9876543210"])
    ws.append(["23AIDS002", "Student Two", "9876501234"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_marksheet_template(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Marksheet"
    ws.append([
        "Reg No",
        "Student Name",
        "Department",
        "Attendance",
        "Subject 1",
        "Subject 2",
        "Subject 3",
        "No of Subjects Failed",
        "GPA",
    ])
    ws.append(["23AIDS001", "Student One", "AIDS", "92%", "88", "91", "84", "0", "8.74"])
    ws.append(["23AIDS002", "Student Two", "AIDS", "89%", "79", "86", "90", "1", "8.11"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    project_root = Path(__file__).resolve().parents[1]
    static_assets = project_root / "frontend" / "static" / "assets"
    static_assets.mkdir(parents=True, exist_ok=True)

    docs = {
        "doc_admin.pdf": (
            "IIT-G Parent Connect - Admin Documentation",
            "System Admin",
            [
                ("Core Responsibilities", [
                    "Create and maintain user accounts with correct role permissions.",
                    "Upload marksheets and supervise department-wide workflows.",
                    "Monitor message logs, session activity, and data consistency.",
                ]),
                ("Daily Workflow", [
                    "Open the Users tab and validate counselor mappings.",
                    "Use Departments and Reports tabs for upload and control.",
                    "Review delivery status and corrective actions in Message Logs.",
                ]),
                ("Operational Safeguards", [
                    "Keep department access state updated before test dispatch.",
                    "Rotate passwords for inactive or compromised accounts.",
                    "Use student template files to maintain import consistency.",
                ]),
            ],
        ),
        "doc_chief_admin.pdf": (
            "IIT-G Parent Connect - Chief Admin Documentation",
            "Chief Admin",
            [
                ("Scope-Based Controls", [
                    "Manage counselors only inside assigned department/year scope.",
                    "Validate test metadata before enabling message send workflows.",
                    "Use scoped credential reset for counselor access recovery.",
                ]),
                ("Data Supervision", [
                    "Ensure student allocations align with counselor department.",
                    "Verify marks parsing and test-level status before dispatch.",
                    "Coordinate with system admin for blocked department actions.",
                ]),
                ("Best Practices", [
                    "Use consistent naming for tests, semesters, and batches.",
                    "Confirm phone number presence before send operations.",
                    "Track progress using filtered activity and reports.",
                ]),
            ],
        ),
        "doc_counsellor.pdf": (
            "IIT-G Parent Connect - Counselor Documentation",
            "Counselor",
            [
                ("Message Sending Flow", [
                    "Open Test Database and choose the required test.",
                    "Use Send Results to preview and customize message layout.",
                    "Arrange subject order and metrics before WhatsApp dispatch.",
                ]),
                ("Student Data Hygiene", [
                    "Maintain accurate student register numbers and parent phone.",
                    "Use provided student template for clean uploads.",
                    "Keep department-specific records updated for each batch.",
                ]),
                ("Troubleshooting", [
                    "If a phone is missing, update student phone and retry.",
                    "If test is blocked, contact admin/chief admin for access.",
                    "Review message history after each send operation.",
                ]),
            ],
        ),
    }

    for filename, payload in docs.items():
        title, audience, sections = payload
        doc_path = static_assets / filename
        make_pdf(doc_path, title, audience, sections)

    student_template = static_assets / "student_list.xlsx"
    marksheet_template = static_assets / "marksheet.xlsx"
    make_student_template(student_template)
    make_marksheet_template(marksheet_template)

    print("Generated documentation PDFs and Excel templates in frontend/static/assets")


if __name__ == "__main__":
    main()
