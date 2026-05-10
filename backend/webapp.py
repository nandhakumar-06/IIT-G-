# webapp.py - Flask Web Application for IIT-G Parent Connect
"""
Complete Flask web application replacing the Streamlit UI.
Serves HTML templates with a dark glass-morphism theme.

Legacy note: this file is retained for compatibility only.
Use backend/app.py as the canonical application entrypoint.
"""
import os
import io
import csv
import json
import uuid
from datetime import datetime
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, send_file, Response
)
from fpdf import FPDF

import database as db
from config import (
    SECRET_KEY, APP_NAME, APP_VERSION, DATA_DIR,
    MESSAGE_TEMPLATE, COUNTRY_CODE
)

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.join(os.path.dirname(BASE_DIR), "frontend")

app = Flask(
    __name__,
    template_folder=os.path.join(FRONTEND_DIR, "templates"),
    static_folder=os.path.join(FRONTEND_DIR, "static"),
)
app.secret_key = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB upload limit

db.init_database()


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("login"))
        sid = session.get("session_id", "")
        if not db.validate_session(sid):
            session.clear()
            flash("Session expired. Please log in again.", "error")
            return redirect(url_for("login"))
        db.touch_session(sid)
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Admin access required.", "error")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Context processor – inject common vars into every template
# ---------------------------------------------------------------------------
@app.context_processor
def inject_globals():
    return {
        "app_name": APP_NAME,
        "app_version": APP_VERSION,
        "current_user_email": session.get("user_email"),
        "current_user_name": session.get("user_name"),
        "current_role": session.get("role"),
        "now": datetime.now(),
    }


# ============================= PAGE ROUTES =================================

@app.route("/")
def index():
    if "user_email" in session:
        return redirect(url_for("admin" if session.get("role") == "admin" else "counselor_page"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")

        user = db.authenticate_user(email, password)
        if not user:
            flash("Invalid email or password.", "error")
            return render_template("login.html")

        allowed, msg = db.check_user_access(email)
        if not allowed:
            flash(msg, "error")
            return render_template("login.html")

        sid = str(uuid.uuid4())
        ok, msg = db.register_session(sid, email, request.remote_addr, request.user_agent.string)
        if not ok:
            flash(msg, "error")
            return render_template("login.html")

        session["user_email"] = email
        session["user_name"] = user["name"]
        session["role"] = user["role"]
        session["session_id"] = sid
        session["department"] = user.get("department", "")

        return redirect(url_for("admin" if user["role"] == "admin" else "counselor_page"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    sid = session.get("session_id")
    if sid:
        db.end_session(sid)
    session.clear()
    return redirect(url_for("login"))


# ======================== ADMIN PAGE =======================================

@app.route("/admin")
@login_required
@admin_required
def admin():
    users = db.get_all_users()
    departments = db.get_departments(active_only=False)
    active_sessions = db.get_active_sessions()
    activity = db.get_counselor_activity_summary()
    format_settings = db.get_format_settings()
    messages = db.get_message_history(limit=200)
    msg_stats = db.get_message_stats()

    counselors = [u for u in users if u["role"] == "counselor"]
    return render_template(
        "admin.html",
        users=users,
        departments=departments,
        sessions=active_sessions,
        activity=activity,
        format_settings=format_settings,
        messages=messages,
        msg_stats=msg_stats,
        counselor_count=len(counselors),
        active_counselor_count=sum(1 for c in counselors if c["is_active"]),
        session_count=len(active_sessions),
    )


# ======================== COUNSELOR PAGE ===================================

@app.route("/counselor")
@login_required
def counselor_page():
    email = session["user_email"]
    user = db.get_user(email)
    if not user:
        session.clear()
        return redirect(url_for("login"))

    students = db.get_students(email)
    tests = db.get_tests_by_counselor(email)
    msg_stats = db.get_message_stats(email)
    msg_history = db.get_message_history(email, limit=50)

    return render_template(
        "counselor.html",
        user=user,
        students=students,
        tests=tests,
        msg_stats=msg_stats,
        msg_history=msg_history,
        can_upload_students=bool(user.get("can_upload_students", 1)),
    )


# ============================== API ROUTES ==================================

# ---------- Users -----------------------------------------------------------

@app.route("/api/users", methods=["POST"])
@login_required
@admin_required
def api_create_user():
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "")
    name = request.form.get("name", "").strip()
    role = request.form.get("role", "counselor")
    department = request.form.get("department", "")
    max_students = int(request.form.get("max_students", 30))
    can_upload = request.form.get("can_upload_students") == "on"

    if not email or not password or not name:
        flash("All required fields must be filled.", "error")
        return redirect(url_for("admin"))

    ok, msg = db.create_user(email, password, name, role, department, max_students, can_upload)

    # Optional student file during registration
    if ok and "student_file" in request.files:
        f = request.files["student_file"]
        if f and f.filename:
            try:
                from core.dynamic_parser import parse_student_excel
                parsed = parse_student_excel(f)
                if parsed:
                    added = db.add_students_bulk(email, parsed)
                    flash(f"User created — {added} students uploaded.", "success")
                else:
                    flash("User created but no valid students found in the uploaded file.", "warning")
            except Exception as e:
                flash(f"User created but student upload failed: {e}", "warning")
            return redirect(url_for("admin"))

    flash(msg, "success" if ok else "error")
    return redirect(url_for("admin"))


@app.route("/api/users/<path:email>/update", methods=["POST"])
@login_required
@admin_required
def api_update_user(email):
    updates = {}
    for key in ("name", "department"):
        val = request.form.get(key, "").strip()
        if val:
            updates[key] = val
    pw = request.form.get("password", "").strip()
    if pw:
        updates["password"] = pw
    ms = request.form.get("max_students")
    if ms:
        updates["max_students"] = int(ms)
    updates["can_upload_students"] = 1 if request.form.get("can_upload_students") == "on" else 0

    db.update_user(email, **updates)
    flash("User updated.", "success")
    return redirect(url_for("admin"))


@app.route("/api/users/<path:email>/delete", methods=["POST"])
@login_required
@admin_required
def api_delete_user(email):
    db.delete_user(email)
    flash("User deleted.", "success")
    return redirect(url_for("admin"))


@app.route("/api/users/<path:email>/lock", methods=["POST"])
@login_required
@admin_required
def api_lock_user(email):
    db.lock_user(email, request.form.get("reason", "Locked by admin"))
    flash("User locked.", "success")
    return redirect(url_for("admin"))


@app.route("/api/users/<path:email>/unlock", methods=["POST"])
@login_required
@admin_required
def api_unlock_user(email):
    db.unlock_user(email)
    flash("User unlocked.", "success")
    return redirect(url_for("admin"))


@app.route("/api/users/<path:email>/upload-students", methods=["POST"])
@login_required
@admin_required
def api_upload_students_for_counselor(email):
    f = request.files.get("student_file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return redirect(url_for("admin"))
    try:
        from core.dynamic_parser import parse_student_excel
        parsed = parse_student_excel(f)
        if parsed:
            added = db.add_students_bulk(email, parsed)
            flash(f"{added} students uploaded for {email}.", "success")
        else:
            flash("No valid students found.", "error")
    except Exception as e:
        flash(f"Upload failed: {e}", "error")
    return redirect(url_for("admin"))


@app.route("/api/users/<path:email>/force-logout", methods=["POST"])
@login_required
@admin_required
def api_force_logout(email):
    db.force_logout_user(email, "admin_action")
    flash(f"Force-logged-out {email}.", "success")
    return redirect(url_for("admin"))


# ---------- Departments -----------------------------------------------------

@app.route("/api/departments", methods=["POST"])
@login_required
@admin_required
def api_create_department():
    code = request.form.get("code", "").strip().upper()
    name = request.form.get("name", "").strip()
    color = request.form.get("color", "#667eea")
    if not code or not name:
        flash("Code and name are required.", "error")
        return redirect(url_for("admin"))
    ok, msg = db.create_department(code, name, color)
    flash(msg, "success" if ok else "error")
    return redirect(url_for("admin"))


@app.route("/api/departments/<int:dept_id>/delete", methods=["POST"])
@login_required
@admin_required
def api_delete_department(dept_id):
    db.delete_department(dept_id)
    flash("Department deleted.", "success")
    return redirect(url_for("admin"))


@app.route("/api/departments/<int:dept_id>/toggle", methods=["POST"])
@login_required
@admin_required
def api_toggle_department(dept_id):
    is_active = request.form.get("is_active") == "1"
    db.update_department(dept_id, is_active=0 if is_active else 1)
    flash("Department updated.", "success")
    return redirect(url_for("admin"))


# ---------- Sessions --------------------------------------------------------

@app.route("/api/sessions/cleanup", methods=["POST"])
@login_required
@admin_required
def api_cleanup_sessions():
    db.cleanup_stale_sessions()
    db.clear_inactive_sessions()
    flash("Sessions cleaned.", "success")
    return redirect(url_for("admin"))


@app.route("/api/sessions/logout-all", methods=["POST"])
@login_required
@admin_required
def api_logout_all():
    db.logout_all_users()
    flash("All users logged out.", "success")
    return redirect(url_for("admin"))


# ---------- Activity Export -------------------------------------------------

@app.route("/api/activity/export/csv")
@login_required
@admin_required
def api_export_activity_csv():
    data = db.get_counselor_activity_summary()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Name", "Email", "Department", "Students", "Tests",
                "Messages", "Status", "Last Login"])
    for a in data:
        w.writerow([a["name"], a["email"], a["department"],
                     a["student_count"], a["tests_uploaded"],
                     a["total_messages"], a["work_status"],
                     a["last_login"] or "Never"])
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=counselor_activity.csv"},
    )


@app.route("/api/activity/export/excel")
@login_required
@admin_required
def api_export_activity_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    data = db.get_counselor_activity_summary()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Counselor Activity"

    headers = ["Name", "Email", "Department", "Students", "Tests",
               "Messages", "Week Msgs", "Status", "Last Login"]
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, a in enumerate(data, 2):
        ws.cell(row=ri, column=1, value=a["name"])
        ws.cell(row=ri, column=2, value=a["email"])
        ws.cell(row=ri, column=3, value=a["department"])
        ws.cell(row=ri, column=4, value=a["student_count"])
        ws.cell(row=ri, column=5, value=a["tests_uploaded"])
        ws.cell(row=ri, column=6, value=a["total_messages"])
        ws.cell(row=ri, column=7, value=a["week_messages"])
        ws.cell(row=ri, column=8, value=a["work_status"])
        ws.cell(row=ri, column=9, value=a["last_login"] or "Never")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="counselor_activity.xlsx")


@app.route("/api/activity/export/pdf")
@login_required
@admin_required
def api_export_activity_pdf():
    data = db.get_counselor_activity_summary()
    pdf = FPDF("L")
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 12, "IIT-G Parent Connect - Counselor Activity", 0, 1, "C")
    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}", 0, 1, "C")
    pdf.ln(4)

    widths = [45, 55, 28, 22, 20, 26, 42, 40]
    heads = ["Name", "Email", "Dept", "Students", "Tests", "Messages", "Status", "Last Login"]
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(102, 126, 234)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(widths, heads):
        pdf.cell(w, 8, h, 1, 0, "C", True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 8)
    for a in data:
        login = (a["last_login"] or "Never")[:16]
        pdf.cell(widths[0], 7, a["name"][:22], 1)
        pdf.cell(widths[1], 7, a["email"][:28], 1)
        pdf.cell(widths[2], 7, a["department"][:10], 1, 0, "C")
        pdf.cell(widths[3], 7, str(a["student_count"]), 1, 0, "C")
        pdf.cell(widths[4], 7, str(a["tests_uploaded"]), 1, 0, "C")
        pdf.cell(widths[5], 7, str(a["total_messages"]), 1, 0, "C")
        pdf.cell(widths[6], 7, a["work_status"][:22], 1, 0, "C")
        pdf.cell(widths[7], 7, login, 1, 1)

    buf = io.BytesIO()
    raw = pdf.output(dest="S")
    buf.write(raw if isinstance(raw, bytes) else raw.encode("latin-1"))
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name="counselor_activity.pdf")


@app.route("/api/activity/<path:email>")
@login_required
@admin_required
def api_activity_detail(email):
    detail = db.get_counselor_detailed_activity(email)
    if not detail:
        return jsonify({"error": "Not found"}), 404
    # Make JSON-serializable
    for key in list(detail.keys()):
        val = detail[key]
        if isinstance(val, datetime):
            detail[key] = val.isoformat()
    return jsonify(detail)


# ---------- Counselor student / marks / reports -----------------------------

@app.route("/api/students/upload", methods=["POST"])
@login_required
def api_upload_students():
    email = session["user_email"]
    user = db.get_user(email)
    if not user.get("can_upload_students", 1):
        flash("You do not have permission to upload students.", "error")
        return redirect(url_for("counselor_page"))

    f = request.files.get("student_file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return redirect(url_for("counselor_page"))

    try:
        from core.dynamic_parser import parse_student_excel
        parsed = parse_student_excel(f)
        if parsed:
            added = db.add_students_bulk(email, parsed)
            flash(f"{added} students uploaded successfully.", "success")
        else:
            flash("No valid students found in the file.", "error")
    except Exception as e:
        flash(f"Upload failed: {e}", "error")
    return redirect(url_for("counselor_page"))


@app.route("/api/students/<reg_no>/delete", methods=["POST"])
@login_required
def api_delete_student(reg_no):
    db.delete_student(session["user_email"], reg_no)
    flash("Student removed.", "success")
    return redirect(url_for("counselor_page"))


@app.route("/api/students/delete-all", methods=["POST"])
@login_required
def api_delete_all_students():
    db.delete_all_students(session["user_email"])
    flash("All students removed.", "success")
    return redirect(url_for("counselor_page"))


@app.route("/api/marks/upload", methods=["POST"])
@login_required
def api_upload_marks():
    email = session["user_email"]
    f = request.files.get("marks_file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return redirect(url_for("counselor_page"))

    try:
        from core.intelligent_parser import IntelligentParser
        parser = IntelligentParser()
        test_info, students = parser.parse_file(f, f.filename)

        if not students:
            flash("No student marks data found in file.", "error")
            return redirect(url_for("counselor_page"))

        subjects = [s["name"] for s in test_info.subjects]
        student_data = [s.to_dict() for s in students]

        ok, msg = db.save_test_marks(
            test_info.test_name or "Unit Test",
            test_info.semester or "1",
            email,
            student_data,
            subjects,
        )
        if ok:
            flash(f"Marks uploaded — {len(students)} students, {len(subjects)} subjects.", "success")
        else:
            flash(f"Error: {msg}", "error")
    except Exception as e:
        flash(f"Parse error: {e}", "error")
    return redirect(url_for("counselor_page"))


@app.route("/api/reports/generate", methods=["POST"])
@login_required
def api_generate_reports():
    email = session["user_email"]
    test_id = request.form.get("test_id")
    reg_nos = request.form.getlist("reg_nos")
    fmt = request.form.get("format", "message")

    if not test_id or not reg_nos:
        flash("Select a test and at least one student.", "error")
        return redirect(url_for("counselor_page"))

    test_id = int(test_id)
    user = db.get_user(email)
    test_meta = db.get_test_metadata(test_id)
    students = db.get_students(email)
    lookup = {s["reg_no"]: s for s in students}

    from utils.whatsapp_helper import get_whatsapp_link
    from utils.template_engine import TemplateEngine

    reports = []
    for rn in reg_nos:
        marks = db.get_student_marks_for_reg(test_id, rn)
        stu = lookup.get(rn, {})
        if not marks:
            continue

        marks_table = TemplateEngine.format_marks_table_simple(marks)
        msg = TemplateEngine.fill_template(
            MESSAGE_TEMPLATE,
            app_name=APP_NAME,
            reg_no=rn,
            student_name=stu.get("student_name", rn),
            department=stu.get("department", ""),
            test_name=(test_meta or {}).get("test_name", "Unit Test"),
            subjects_table=marks_table,
            counselor_name=user["name"],
        )

        phone = stu.get("parent_phone", "")
        wa = get_whatsapp_link(phone, msg) if phone else ""
        db.log_message(email, rn, stu.get("student_name", ""), msg, fmt, wa)

        reports.append({
            "reg_no": rn,
            "name": stu.get("student_name", rn),
            "marks": marks,
            "message": msg,
            "whatsapp_link": wa,
            "phone": phone,
        })

    session["reports"] = reports
    session["report_test_id"] = test_id
    flash(f"Reports generated for {len(reports)} students.", "success")
    return redirect(url_for("counselor_page"))


@app.route("/api/reports/pdf/<reg_no>")
@login_required
def api_student_pdf(reg_no):
    email = session["user_email"]
    test_id = request.args.get("test_id")
    if not test_id:
        return "Missing test_id", 400

    marks = db.get_student_marks_for_reg(int(test_id), reg_no)
    if not marks:
        return "No marks found", 404

    stu = next((s for s in db.get_students(email) if s["reg_no"] == reg_no), None)
    user = db.get_user(email)
    meta = db.get_test_metadata(int(test_id))

    from utils.pdf_generator import generate_student_pdf
    pdf_bytes = generate_student_pdf(
        student_name=stu["student_name"] if stu else reg_no,
        reg_no=reg_no,
        department=(stu or {}).get("department", ""),
        subjects_marks=marks,
        counselor_name=user["name"],
        test_name=(meta or {}).get("test_name", "Unit Test"),
    )
    return send_file(
        io.BytesIO(pdf_bytes), mimetype="application/pdf",
        as_attachment=True, download_name=f"{reg_no}_report.pdf",
    )


# ---------- Format Settings -------------------------------------------------

@app.route("/api/settings/format", methods=["POST"])
@login_required
@admin_required
def api_update_format_settings():
    default = request.form.get("default_format", "message")
    allowed = request.form.getlist("allowed_formats")
    bulk = request.form.get("bulk_format", "same_as_individual")
    db.update_format_settings(default, allowed, bulk, session["user_email"])
    flash("Format settings updated.", "success")
    return redirect(url_for("admin"))


# ---------- Static data assets ----------------------------------------------

@app.route("/data/<path:filename>")
def serve_data(filename):
    return send_file(os.path.join(DATA_DIR, filename))


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    from app import app as canonical_app

    port = int(sys.argv[1]) if len(sys.argv) > 1 else 5000
    print("\n  [INFO] webapp.py is legacy. Delegating launch to backend/app.py\n")
    print(f"  ✅  IIT-G Parent Connect running at: http://localhost:{port}\n")
    canonical_app.run(debug=False, host="0.0.0.0", port=port, use_reloader=False)
