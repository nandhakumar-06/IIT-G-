# webapp.py - Flask Web Application for RMKCET Parent Connect
"""
Complete Flask web application replacing the Streamlit UI.
Serves HTML templates with a dark glass-morphism theme.
"""
import os
import io
import csv
import json
import hashlib
import shutil
import sqlite3
import smtplib
import uuid
from datetime import datetime, timedelta
from functools import wraps
from urllib.parse import urlparse, parse_qs
import re

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, send_file, Response, abort
)
from fpdf import FPDF

import database as db
from config import (
    SECRET_KEY, APP_NAME, APP_VERSION, DATA_DIR, DATABASE_FILE,
    MESSAGE_TEMPLATE, COUNTRY_CODE, DEPT_REG_PATTERNS, OTP_EXPIRY_SECONDS,
    SMTP_SERVER, SMTP_PORT, SMTP_USERNAME, SMTP_PASSWORD, TEST_MODE,
    DEFAULT_ADMIN
)
from models.test_metadata import normalize_test_name
from utils.email_helper import send_email
from utils.otp_helper import generate_otp

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.join(os.path.dirname(BASE_DIR), "frontend")
STATIC_ASSETS_DIR = os.path.join(FRONTEND_DIR, "static", "assets")

app = Flask(
    __name__,
    template_folder=os.path.join(FRONTEND_DIR, "templates"),
    static_folder=os.path.join(FRONTEND_DIR, "static"),
)
app.secret_key = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB upload limit

db.init_database()

ALLOWED_TEST_NAMES = {"IAT 1", "IAT 2", "MODEL EXAM"}
_SMTP_STATUS_CACHE = {"checked_at": None, "result": None}


# ---------------------------------------------------------------------------
# Auth helpers - Industrial-grade session validation
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("login"))
        
        sid = session.get("session_id", "")
        
        # Use industrial-grade session validation
        is_valid, reason, user_email = db.validate_session_strict(
            sid, 
            request.remote_addr, 
            request.user_agent.string
        )
        
        if not is_valid:
            session.clear()
            # Provide user-friendly messages based on reason
            if "session_timeout" in reason:
                flash("Your session has expired due to inactivity. Please log in again.", "warning")
            elif "session_inactive" in reason:
                if "new_login" in reason or "new_device" in reason:
                    flash("You have been logged out because you logged in from another device.", "warning")
                elif "admin_action" in reason:
                    flash("An administrator has logged you out.", "warning")
                else:
                    flash("Your session is no longer valid. Please log in again.", "error")
            elif reason == "user_deactivated":
                flash("Your account has been deactivated. Contact an administrator.", "error")
            elif reason == "user_locked":
                flash("Your account is locked. Contact an administrator.", "error")
            else:
                flash("Session expired. Please log in again.", "error")
            return redirect(url_for("login"))
        
        _maybe_flash_disabled_department_warning()

        # Update session activity (heartbeat)
        db.touch_session(sid)
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") not in {"admin", "chief_admin", "hod", "deo", "principal"}:
            flash("Admin access required.", "error")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


def _is_system_admin(role):
    return str(role or "").strip().lower() == "admin"


def _is_chief_admin(role):
    return str(role or "").strip().lower() == "chief_admin"


def _is_hod(role):
    return str(role or "").strip().lower() in {"chief_admin", "hod"}


def _is_deo(role):
    return str(role or "").strip().lower() == "deo"


def _is_principal(role):
    return str(role or "").strip().lower() == "principal"


def _is_counselor(role):
    return str(role or "").strip().lower() == "counselor"


def _is_admin_portal_user(role):
    return _is_system_admin(role) or _is_hod(role) or _is_deo(role) or _is_principal(role)


def _normalize_allowed_test_name(value):
    normalized = normalize_test_name(value)
    if normalized in ALLOWED_TEST_NAMES:
        return normalized
    return ""


def _allowed_tabs_for_role(role):
    if _is_system_admin(role):
        return {"users", "departments", "reports", "activity", "monitoring", "messages", "config", "database"}
    if _is_hod(role):
        return {"dashboard", "reports", "activity", "messages"}
    if _is_deo(role):
        return {"reports", "activity", "users", "messages"}
    if _is_principal(role):
        return {"dashboard", "reports", "activity", "users", "database"}
    return set()


def _default_tab_for_role(role):
    if _is_system_admin(role):
        return "reports"
    if _is_principal(role):
        return "dashboard"
    if _is_hod(role):
        return "dashboard"
    if _is_deo(role):
        return "reports"
    return "recent-tests"


def _mask_email(email):
    value = str(email or "").strip()
    if "@" not in value:
        return value
    local, domain = value.split("@", 1)
    if len(local) <= 2:
        local_masked = local[0] + "*" if local else "*"
    else:
        local_masked = local[0] + ("*" * (len(local) - 2)) + local[-1]
    return f"{local_masked}@{domain}"


def _is_truthy(value):
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _get_security_config_flags():
    cfg = db.get_app_config() or {}
    return {
        "require_otp_on_password_reset": _is_truthy(cfg.get("require_otp_on_password_reset", "false")),
        "require_otp_on_login": _is_truthy(cfg.get("require_otp_on_login", "false")),
        "disable_default_admin_on_new_system_admin": _is_truthy(cfg.get("disable_default_admin_on_new_system_admin", "false")),
    }


def _get_default_admin_email():
    return str((DEFAULT_ADMIN or {}).get("email") or "").strip().lower()


def _is_default_admin_user(email, role=None):
    if role is not None and not _is_system_admin(role):
        return False
    return str(email or "").strip().lower() == _get_default_admin_email()


def _is_login_otp_required_for_user(role, email):
    flags = _get_security_config_flags()
    return bool(flags.get("require_otp_on_login") and not _is_default_admin_user(email, role))


def _is_password_reset_otp_required_for_user(role, email):
    flags = _get_security_config_flags()
    return bool(flags.get("require_otp_on_password_reset") and not _is_default_admin_user(email, role))


def _send_otp_email(target_email, otp_code, purpose):
    subject = f"RMKCET SHINE - {purpose} OTP"
    body = (
        f"<h3>{purpose} verification</h3>"
        f"<p>Your one-time password is:</p>"
        f"<h2 style='letter-spacing:4px;color:#2563eb'>{otp_code}</h2>"
        f"<p>This OTP is valid for {max(1, int(OTP_EXPIRY_SECONDS // 60))} minutes.</p>"
        f"<p>If you did not initiate this request, you can ignore this mail.</p>"
        f"<p>RMKCET SHINE</p>"
    )
    return send_email(target_email, subject, body, html=True)


def _resolve_smtp_status(force_refresh=False):
    now = datetime.now()
    cached_at = _SMTP_STATUS_CACHE.get("checked_at")
    if (
        not force_refresh
        and cached_at is not None
        and (now - cached_at).total_seconds() < 180
        and _SMTP_STATUS_CACHE.get("result")
    ):
        return dict(_SMTP_STATUS_CACHE["result"])

    if TEST_MODE:
        result = {
            "state": "test",
            "label": "SMTP Test Mode",
            "icon": "fa-flask",
            "detail": "TEST_MODE is enabled. SMTP delivery is bypassed.",
        }
    elif not SMTP_USERNAME or not SMTP_PASSWORD:
        result = {
            "state": "missing",
            "label": "SMTP Missing",
            "icon": "fa-triangle-exclamation",
            "detail": "SMTP username/password are not configured.",
        }
    else:
        try:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=6) as server:
                server.ehlo()
                server.starttls()
                server.ehlo()
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
            result = {
                "state": "ready",
                "label": "SMTP Ready",
                "icon": "fa-circle-check",
                "detail": f"Connected to {SMTP_SERVER}:{SMTP_PORT} as {SMTP_USERNAME}",
            }
        except Exception as exc:
            result = {
                "state": "error",
                "label": "SMTP Error",
                "icon": "fa-circle-xmark",
                "detail": f"{type(exc).__name__}: {exc}",
            }

    _SMTP_STATUS_CACHE["checked_at"] = now
    _SMTP_STATUS_CACHE["result"] = dict(result)
    return result


def _get_cached_smtp_status():
    cached = _SMTP_STATUS_CACHE.get("result")
    if cached:
        return dict(cached)
    if TEST_MODE:
        return {
            "state": "test",
            "label": "SMTP Test Mode",
            "icon": "fa-flask",
            "detail": "TEST_MODE is enabled. SMTP delivery is bypassed.",
        }
    if not SMTP_USERNAME or not SMTP_PASSWORD:
        return {
            "state": "missing",
            "label": "SMTP Missing",
            "icon": "fa-triangle-exclamation",
            "detail": "SMTP username/password are not configured.",
        }
    return {
        "state": "unknown",
        "label": "SMTP Pending",
        "icon": "fa-clock",
        "detail": "Status check runs in background after login.",
    }


def _panel_endpoint_for_role(role):
    if _is_principal(role):
        return "principal_panel"
    if _is_hod(role):
        return "hod_panel"
    if _is_deo(role):
        return "deo_panel"
    return "admin"


def _tutorial_role_key(role):
    role_norm = str(role or "").strip().lower()
    if role_norm in {"chief_admin", "hod"}:
        return "hod"
    return role_norm


def _tutorial_flags_from_config(config):
    cfg = config or {}
    return {
        "master": str(cfg.get("tutorial_master_enabled", "true")).strip().lower() == "true",
        "counselor": str(cfg.get("tutorial_counselor_enabled", "true")).strip().lower() == "true",
        "hod": str(cfg.get("tutorial_hod_enabled", "true")).strip().lower() == "true",
        "deo": str(cfg.get("tutorial_deo_enabled", "true")).strip().lower() == "true",
        "principal": str(cfg.get("tutorial_principal_enabled", "true")).strip().lower() == "true",
    }


def _is_tutorial_enabled_for_role(role, config=None):
    role_key = _tutorial_role_key(role)
    if role_key not in {"counselor", "hod", "deo", "principal"}:
        return False

    flags = _tutorial_flags_from_config(config or db.get_app_config())
    return bool(flags.get("master") and flags.get(role_key))


def _format_year_level(value):
    """Render year values as 1st Year, 2nd Year, etc."""
    raw = str(value or "").strip()
    if not raw:
        return ""

    try:
        year = int(raw)
    except (TypeError, ValueError):
        return raw

    suffix = "th"
    if year % 100 not in (11, 12, 13):
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(year % 10, "th")
    return f"{year}{suffix} Year"


def _get_actor_scope_pairs(actor_email, actor_role):
    if _is_system_admin(actor_role):
        return None
    if _is_principal(actor_role):
        return None
    scopes = db.get_chief_admin_scopes(actor_email)
    return {(str(s.get("department") or "").upper(), int(s.get("year_level") or 1)) for s in scopes}


def _parse_scope_pairs_from_form(form_key="scope_pairs"):
    """Parse repeated scope inputs encoded as DEPT::YEAR pairs."""
    parsed = []
    for raw in request.form.getlist(form_key):
        val = str(raw or "").strip()
        if "::" not in val:
            continue
        dep_part, year_part = val.split("::", 1)
        dep = dep_part.strip().upper()
        try:
            year_level = int(str(year_part).strip())
        except (TypeError, ValueError):
            continue
        if dep and year_level in (1, 2, 3, 4):
            parsed.append((dep, year_level))
    return sorted(set(parsed))


def _scope_pairs_to_nav_map(scope_pairs):
    """Convert scope tuple set/list into sorted dept list and dept->years map."""
    years_by_department = {}
    for dep, yr in sorted(set(scope_pairs or [])):
        years_by_department.setdefault(dep, []).append(int(yr))
    for dep in list(years_by_department.keys()):
        years_by_department[dep] = sorted(set(years_by_department[dep]))
    departments = sorted(years_by_department.keys())
    return departments, years_by_department


def _build_example_scope_pairs(limit=4):
    """Build deterministic example scopes for system-admin test mode preview."""
    active_departments = db.get_departments(active_only=True)
    pairs = []
    for d in active_departments:
        dep = str(d.get("code") or "").strip().upper()
        if not dep:
            continue
        pairs.append((dep, 1))
        pairs.append((dep, 2))
        if len(pairs) >= limit:
            break
    return sorted(set(pairs[:limit]))


def _get_assigned_departments_for_user(user_email, user_role):
    """Return distinct department codes assigned to counselor/HoD/DEO users."""
    role_norm = str(user_role or "").strip().lower()
    email = str(user_email or "").strip()
    if not email:
        return []

    if role_norm in {"counselor"}:
        user = db.get_user(email) or {}
        dep = str(user.get("department") or "").strip().upper()
        return [dep] if dep else []

    if role_norm in {"chief_admin", "hod", "deo"}:
        scope_rows = db.get_chief_admin_scopes(email)
        deps = {
            str(s.get("department") or "").strip().upper()
            for s in scope_rows
            if s.get("department")
        }
        return sorted([d for d in deps if d])

    return []


def _build_assigned_department_badges(user_email, user_role):
    """Build department badges with active/disabled flags for sidebar display."""
    department_codes = _get_assigned_departments_for_user(user_email, user_role)
    if not department_codes:
        return []

    all_departments = db.get_departments(active_only=False)
    active_map = {
        str(d.get("code") or "").strip().upper(): bool(d.get("is_active"))
        for d in all_departments
        if d.get("code")
    }
    return [
        {
            "code": dep,
            "is_active": active_map.get(dep, True),
        }
        for dep in department_codes
    ]


def _maybe_flash_disabled_department_warning():
    """Show a one-time warning per disabled assigned department for non-principal roles."""
    role = session.get("role")
    if _is_principal(role) or _is_system_admin(role):
        return

    email = session.get("user_email")
    assigned_departments = _get_assigned_departments_for_user(email, role)
    if not assigned_departments:
        return

    disabled_departments = [d for d in assigned_departments if not db.is_department_active(d)]
    if not disabled_departments:
        return

    warned = {str(x).strip().upper() for x in (session.get("disabled_department_warning_seen") or [])}
    pending = [d for d in disabled_departments if d not in warned]
    if not pending:
        return

    flash(
        "Warning: Your assigned department(s) are currently disabled: " + ", ".join(sorted(pending)),
        "warning",
    )
    session["disabled_department_warning_seen"] = sorted(warned | set(pending))
    session.modified = True


def _can_chief_admin_touch_user(actor_email, target_user):
    if not target_user:
        return False
    
    target_role = target_user.get("role")
    
    # HoD can manage counselors/DEOs in their scope.
    if target_role in {"counselor", "deo"}:
        scopes = _get_actor_scope_pairs(actor_email, "hod") or set()
        key = (str(target_user.get("department") or "").upper(), int(target_user.get("year_level") or 1))
        return key in scopes
    
    # HoD can manage other HoD accounts only with scope overlap.
    if target_role in {"chief_admin", "hod"}:
        actor_scopes = _get_actor_scope_pairs(actor_email, "hod") or set()
        target_scopes = db.get_chief_admin_scopes(target_user.get("email"))
        target_scopes_set = {(str(s.get("department") or "").upper(), int(s.get("year_level") or 1)) for s in target_scopes}
        return bool(actor_scopes & target_scopes_set)  # Check for scope intersection
    
    return False


def _can_deo_touch_user(actor_email, target_user):
    if not target_user:
        return False
    if str(target_user.get("role") or "").strip().lower() != "counselor":
        return False
    scopes = _get_actor_scope_pairs(actor_email, "deo") or set()
    key = (str(target_user.get("department") or "").upper(), int(target_user.get("year_level") or 1))
    return key in scopes


def _get_allowed_counselor_emails_for_actor(actor_email, actor_role, forced_scope_pairs=None):
    if _is_system_admin(actor_role):
        return None
    users = db.get_scoped_users_for_admin(actor_email, actor_role)
    if forced_scope_pairs is not None and actor_role in {"hod", "deo"}:
        forced_scope_set = {(str(dep).upper(), int(yr)) for dep, yr in forced_scope_pairs}
        all_users = db.get_all_users()
        preview_users = []
        for u in all_users:
            role_name = str(u.get("role") or "").strip().lower()
            if role_name in {"counselor", "deo"}:
                key = (str(u.get("department") or "").strip().upper(), int(u.get("year_level") or 1))
                if key in forced_scope_set:
                    preview_users.append(u)
                continue
            if role_name in {"chief_admin", "hod"}:
                other_scopes = {
                    (str(s.get("department") or "").strip().upper(), int(s.get("year_level") or 1))
                    for s in db.get_chief_admin_scopes(u.get("email"))
                }
                if forced_scope_set & other_scopes:
                    preview_users.append(u)
                continue
        users = preview_users
    return [u.get("email") for u in users if u.get("role") == "counselor"]


def _can_manage_department_year(actor_email, actor_role, department, year_level):
    if _is_system_admin(actor_role):
        return True
    if _is_deo(actor_role):
        scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
        return (str(department or "").strip().upper(), int(year_level or 1)) in scopes
    if not _is_hod(actor_role):
        return False
    scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
    return (str(department or "").strip().upper(), int(year_level or 1)) in scopes


def _is_counselor_department_blocked(user_email, role):
    if str(role or "") != "counselor":
        return False
    user = db.get_user(user_email) or {}
    return not db.is_department_active(user.get("department"))


def _filter_activity_for_actor(activity_rows, actor_email, actor_role):
    scopes = _get_actor_scope_pairs(actor_email, actor_role)
    if scopes is None:
        return activity_rows
    filtered = []
    for row in activity_rows:
        u = db.get_user(row.get("email")) or {}
        key = (str(row.get("department") or "").upper(), int(u.get("year_level") or 1))
        if key in scopes:
            filtered.append(row)
    return filtered


def _ensure_system_admin(default_tab="users"):
    if _is_system_admin(session.get("role")):
        return None
    flash("Only system admin can perform this action.", "error")
    return _redirect_admin_back(default_tab)


def _get_admin_tab(default_tab="users"):
    """Resolve active admin tab from form/query/referrer."""
    tab = (request.form.get("tab") or request.args.get("tab") or "").strip()
    if tab:
        return tab

    ref = (request.referrer or "").strip()
    if ref:
        try:
            parsed = urlparse(ref)
            ref_tab = parse_qs(parsed.query).get("tab", [""])[0].strip()
            if ref_tab:
                return ref_tab
        except Exception:
            pass

    return default_tab


def _redirect_admin_back(default_tab="users", **extra_query):
    params = {"tab": _get_admin_tab(default_tab)}
    for key, value in extra_query.items():
        if value is None:
            continue
        sval = str(value).strip()
        if sval:
            params[key] = sval
    return redirect(url_for(_panel_endpoint_for_role(session.get("role")), **params))


def _get_message_filters_from_request():
    msg_day = (request.values.get("msg_day") or "").strip()
    msg_q = (request.values.get("msg_q") or "").strip()
    msg_year = (request.values.get("msg_year") or "").strip()
    msg_month = (request.values.get("msg_month") or "").strip()
    msg_day_num = (request.values.get("msg_day_num") or "").strip()
    return msg_day, msg_q, msg_year, msg_month, msg_day_num


def _message_export_filename(ext):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"message_activity_{stamp}.{ext}"


def _resolve_asset_file(filename):
    candidate = os.path.join(STATIC_ASSETS_DIR, filename)
    if os.path.isfile(candidate):
        return candidate
    return None


def _parse_mark_as_float(value):
    sval = str(value or "").strip()
    if not sval:
        return None
    up = sval.upper()
    if up in {"ABSENT", "AB", "A", "-", "NA", "N/A"}:
        return None
    try:
        num = float(sval)
    except (TypeError, ValueError):
        return None
    if num < 0 or num > 100:
        return None
    return num


def _build_admin_dashboard_data(actor_email, actor_role, activity_rows, departments):
    scope_pairs = _get_actor_scope_pairs(actor_email, actor_role)
    allowed_dep_year = set(scope_pairs or []) if scope_pairs is not None else None
    allowed_departments = (
        sorted({dep for dep, _ in allowed_dep_year})
        if allowed_dep_year is not None
        else sorted({str(d.get("code") or "").strip().upper() for d in (departments or []) if d.get("code")})
    )

    # 1) Counselor completion across department + year
    completion_by_scope = {}
    completion_by_department = {}
    total_students = 0
    total_reached = 0
    for row in (activity_rows or []):
        dep = str(row.get("department") or "").strip().upper()
        user = db.get_user(row.get("email")) or {}
        year_level = int(user.get("year_level") or 1)
        if allowed_dep_year is not None and (dep, year_level) not in allowed_dep_year:
            continue

        students = int(row.get("student_count") or 0)
        reached = int(row.get("unique_students_messaged") or 0)
        pct = (reached / max(1, students)) * 100 if students else 0

        scope_label = f"{dep} Y{year_level}" if dep else f"Y{year_level}"
        completion_by_scope.setdefault(scope_label, []).append(pct)
        completion_by_department.setdefault(dep or "N/A", {"students": 0, "reached": 0})
        completion_by_department[dep or "N/A"]["students"] += students
        completion_by_department[dep or "N/A"]["reached"] += reached

        total_students += students
        total_reached += reached

    scope_labels = sorted(completion_by_scope.keys())
    scope_values = [round(sum(completion_by_scope[k]) / max(1, len(completion_by_scope[k])), 1) for k in scope_labels]

    dept_completion_labels = sorted(completion_by_department.keys())
    dept_completion_values = []
    for dep in dept_completion_labels:
        dep_students = completion_by_department[dep]["students"]
        dep_reached = completion_by_department[dep]["reached"]
        dept_completion_values.append(round((dep_reached / max(1, dep_students)) * 100 if dep_students else 0, 1))

    overall_completion = round((total_reached / max(1, total_students)) * 100 if total_students else 0, 1)

    # 2) GPA/marks average and 3) test histogram with upload-factor adjustment
    conn = db.get_conn()
    rows = conn.execute(
        """
        SELECT
            UPPER(COALESCE(tm.department, '')) AS department,
            COALESCE(tm.year_level, 1) AS year_level,
            CAST(COALESCE(tm.semester, 0) AS INTEGER) AS semester,
            UPPER(COALESCE(tm.test_name, '')) AS test_name,
            sm.marks AS marks
        FROM test_metadata tm
        JOIN student_marks sm ON sm.test_id = tm.test_id
        """
    ).fetchall()
    conn.close()

    gpa_buckets = {}
    hist_buckets = {}
    for r in rows:
        dep = str(r["department"] or "").strip().upper()
        yr = int(r["year_level"] or 1)
        if allowed_dep_year is not None and (dep, yr) not in allowed_dep_year:
            continue
        mark_val = _parse_mark_as_float(r["marks"])
        if mark_val is None:
            continue

        gpa_key = (dep or "N/A", yr)
        gpa_buckets.setdefault(gpa_key, {"sum": 0.0, "count": 0})
        gpa_buckets[gpa_key]["sum"] += mark_val
        gpa_buckets[gpa_key]["count"] += 1

        sem = int(r["semester"] or 0)
        tname = str(r["test_name"] or "").strip().upper()
        if sem not in (1, 2) or tname not in ALLOWED_TEST_NAMES:
            continue
        hist_key = (sem, tname)
        hist_buckets.setdefault(hist_key, {"sum": 0.0, "count": 0, "departments": set()})
        hist_buckets[hist_key]["sum"] += mark_val
        hist_buckets[hist_key]["count"] += 1
        if dep:
            hist_buckets[hist_key]["departments"].add(dep)

    gpa_labels = []
    gpa_values = []
    for (dep, yr) in sorted(gpa_buckets.keys(), key=lambda x: (x[0], x[1])):
        item = gpa_buckets[(dep, yr)]
        avg_val = item["sum"] / max(1, item["count"])
        gpa_labels.append(f"{dep} Y{yr}")
        gpa_values.append(round(avg_val, 1))

    matrix_labels = []
    matrix_values = []
    matrix_raw_values = []
    matrix_coverage = []
    target_departments = max(1, len(allowed_departments))
    for sem in (1, 2):
        for tname in ("IAT 1", "IAT 2", "MODEL EXAM"):
            key = (sem, tname)
            item = hist_buckets.get(key, {"sum": 0.0, "count": 0, "departments": set()})
            raw_avg = (item["sum"] / max(1, item["count"])) if item["count"] else 0.0
            coverage = len(item["departments"]) / target_departments
            adjusted = raw_avg * coverage

            matrix_labels.append(f"Sem {sem} - {tname}")
            matrix_raw_values.append(round(raw_avg, 1))
            matrix_values.append(round(adjusted, 1))
            matrix_coverage.append(round(coverage * 100, 1))

    return {
        "counselor_activity": {"labels": scope_labels, "values": scope_values},
        "department_gpa": {"labels": gpa_labels, "values": gpa_values},
        "test_histogram": {
            "labels": matrix_labels,
            "values": matrix_values,
            "raw_values": matrix_raw_values,
            "coverage": matrix_coverage,
        },
        "completion_overview": {
            "overall": overall_completion,
            "department_labels": dept_completion_labels,
            "department_values": dept_completion_values,
        },
    }


def _build_counselor_dashboard_data(counselor_email):
    conn = db.get_conn()
    rows = conn.execute(
        """
        SELECT
            cs.reg_no,
            cs.student_name,
            CAST(COALESCE(tm.semester, 0) AS INTEGER) AS semester,
            UPPER(COALESCE(tm.test_name, '')) AS test_name,
            sm.marks AS marks
        FROM counselor_students cs
        LEFT JOIN student_marks sm ON sm.reg_no = cs.reg_no
        LEFT JOIN test_metadata tm ON tm.test_id = sm.test_id
        WHERE cs.counselor_email = ?
        """,
        (counselor_email,),
    ).fetchall()
    conn.close()

    student_totals = {}
    matrix = {}
    all_students = set()
    for r in rows:
        reg = str(r["reg_no"] or "").strip()
        name = str(r["student_name"] or reg).strip()
        if not reg:
            continue
        all_students.add(reg)
        student_totals.setdefault(reg, {"name": name, "sum": 0.0, "count": 0})

        mark_val = _parse_mark_as_float(r["marks"])
        if mark_val is None:
            continue

        student_totals[reg]["sum"] += mark_val
        student_totals[reg]["count"] += 1

        sem = int(r["semester"] or 0)
        tname = str(r["test_name"] or "").strip().upper()
        if sem not in (1, 2) or tname not in ALLOWED_TEST_NAMES:
            continue
        key = (sem, tname)
        matrix.setdefault(key, {"sum": 0.0, "count": 0, "students": set()})
        matrix[key]["sum"] += mark_val
        matrix[key]["count"] += 1
        matrix[key]["students"].add(reg)

    student_labels = []
    student_values = []
    for reg, item in sorted(student_totals.items(), key=lambda x: (x[1]["name"], x[0])):
        if item["count"] <= 0:
            continue
        student_labels.append(item["name"][:24])
        student_values.append(round(item["sum"] / max(1, item["count"]), 1))

    # Keep mobile-friendly chart payload.
    if len(student_labels) > 25:
        student_labels = student_labels[:25]
        student_values = student_values[:25]

    matrix_labels = []
    matrix_values = []
    matrix_raw_values = []
    matrix_coverage = []
    total_students = max(1, len(all_students))
    for sem in (1, 2):
        for tname in ("IAT 1", "IAT 2", "MODEL EXAM"):
            key = (sem, tname)
            item = matrix.get(key, {"sum": 0.0, "count": 0, "students": set()})
            raw_avg = (item["sum"] / max(1, item["count"])) if item["count"] else 0.0
            coverage = len(item["students"]) / total_students
            adjusted = raw_avg * coverage
            matrix_labels.append(f"Sem {sem} - {tname}")
            matrix_raw_values.append(round(raw_avg, 1))
            matrix_values.append(round(adjusted, 1))
            matrix_coverage.append(round(coverage * 100, 1))

    return {
        "student_activity": {"labels": student_labels, "values": student_values},
        "test_histogram": {
            "labels": matrix_labels,
            "values": matrix_values,
            "raw_values": matrix_raw_values,
            "coverage": matrix_coverage,
        },
    }


# ---------------------------------------------------------------------------
# Context processor – inject common vars into every template
# ---------------------------------------------------------------------------
@app.context_processor
def inject_globals():
    current_email = session.get("user_email")
    current_role = session.get("role")
    assigned_departments = _build_assigned_department_badges(current_email, current_role) if current_email else []
    app_config = db.get_app_config()
    tutorial_flags = _tutorial_flags_from_config(app_config)
    tutorial_role_key = _tutorial_role_key(current_role)
    tutorial_enabled_for_current_role = bool(
        tutorial_flags.get("master") and tutorial_flags.get(tutorial_role_key, False)
    )
    security_flags = _get_security_config_flags()
    smtp_status = _get_cached_smtp_status() if _is_system_admin(current_role) else {
        "state": "unknown",
        "label": "SMTP",
        "icon": "fa-envelope",
        "detail": "SMTP status is visible to system admin only.",
    }
    password_reset_otp_for_user = _is_password_reset_otp_required_for_user(current_role, current_email)
    login_otp_for_user = _is_login_otp_required_for_user(current_role, current_email)
    is_default_admin_user = _is_default_admin_user(current_email, current_role)

    return {
        "app_name": APP_NAME,
        "app_version": APP_VERSION,
        "current_user_email": current_email,
        "current_user_name": session.get("user_name"),
        "current_role": current_role,
        "role_panel_endpoint": _panel_endpoint_for_role(current_role),
        "assigned_scope_summary": "",
        "assigned_departments": assigned_departments,
        "tutorial_flags": tutorial_flags,
        "tutorial_role_key": tutorial_role_key,
        "tutorial_enabled_for_current_role": tutorial_enabled_for_current_role,
        "tutorial_auto_welcome": request.args.get("tutorial_welcome") == "1",
        "app_config": app_config,
        "require_otp_on_password_reset": bool(security_flags.get("require_otp_on_password_reset")),
        "require_otp_on_login": bool(security_flags.get("require_otp_on_login")),
        "password_reset_otp_for_user": bool(password_reset_otp_for_user),
        "login_otp_for_user": bool(login_otp_for_user),
        "is_default_admin_user": bool(is_default_admin_user),
        "format_year_level": _format_year_level,
        "smtp_status": smtp_status,
        "now": datetime.now(),
    }


@app.route("/documentation/download", methods=["GET"])
@login_required
def download_role_documentation():
    role = str(session.get("role") or "counselor").strip().lower()
    doc_map = {
        "admin": "doc_admin.pdf",
        "hod": "doc_chief_admin.pdf",
        "chief_admin": "doc_chief_admin.pdf",
        "deo": "doc_counsellor.pdf",
        "principal": "doc_admin.pdf",
        "counselor": "doc_counsellor.pdf",
    }
    filename = doc_map.get(role, "doc_counsellor.pdf")
    target = _resolve_asset_file(filename)
    if not target:
        flash("Documentation file is not available yet.", "error")
        return redirect(request.referrer or url_for("index"))
    return send_file(target, as_attachment=True, download_name=filename)


@app.route("/instructions", methods=["GET"])
@login_required
def instructions_page():
    role = str(session.get("role") or "counselor").strip().lower()
    role_titles = {
        "admin": "System Admin",
        "principal": "Principal",
        "hod": "HoD",
        "chief_admin": "HoD",
        "deo": "DEO",
        "counselor": "Counselor",
    }
    return render_template(
        "instructions.html",
        role_title=role_titles.get(role, "User"),
        panel_endpoint=_panel_endpoint_for_role(role),
        role_key=role,
    )


@app.route("/support/templates/student", methods=["GET"])
@login_required
def download_student_template():
    filename = "student_list.xlsx"
    target = _resolve_asset_file(filename)
    if not target:
        flash("Student template is not available yet.", "error")
        return redirect(request.referrer or url_for("index"))
    return send_file(target, as_attachment=True, download_name=filename)


@app.route("/support/templates/marksheet", methods=["GET"])
@login_required
def download_marksheet_template():
    filename = "marksheet.xlsx"
    target = _resolve_asset_file(filename)
    if not target:
        flash("Marksheet template is not available yet.", "error")
        return redirect(request.referrer or url_for("index"))
    return send_file(target, as_attachment=True, download_name=filename)


def _normalize_metric_key(key):
    return re.sub(r"[^a-z0-9]", "", str(key or "").lower())


def _is_unknown_metric_field(raw_key, key_norm):
    raw = str(raw_key or "").strip().lower()
    if not raw:
        return True
    if raw.startswith("unnamed"):
        return True
    if re.match(r"^subject[_\s-]*\d+$", raw):
        return True
    if key_norm.startswith("unnamed"):
        return True
    return False


def _is_absent_mark(value):
    s = str(value or "").strip().lower()
    return s in {"absent", "ab", "a", "na", "-", "not attended"}


def _build_parent_subjects_table(marks, ordered_fields=None):
    """Build standardized marks block with optional caller-defined ordering."""
    if not isinstance(marks, dict):
        return ""

    attendance_values = []
    gpa_values = []
    failed_values = []
    not_attended_values = []
    subject_rows = []

    attendance_keys = {"attendance", "att"}
    gpa_keys = {"gpa", "cgpa"}
    failed_keys = {"noofsubjectsfailed", "failedsubjects", "failedcount", "nooffailedsubjects"}
    not_attended_keys = {"examnotattended", "notattended", "absentcount", "noofsubjectsabsent"}
    ignored_keys = {
        "regno", "registernumber", "name", "studentname", "department", "section",
        "batch", "semester", "test", "total", "overall", "percentage", "grade",
        "result", "status", "parentphone", "phone", "parentemail", "email",
        "sno", "slno", "serialno", "serialnumber", "rollno",
        "absentees", "absentee", "absentstudents"
    }

    for raw_key, raw_val in marks.items():
        key_norm = _normalize_metric_key(raw_key)
        value = str(raw_val or "").strip()
        if _is_unknown_metric_field(raw_key, key_norm):
            continue
        if not key_norm or key_norm in ignored_keys:
            continue

        if key_norm in attendance_keys:
            attendance_values.append(value)
            continue
        if key_norm in gpa_keys:
            gpa_values.append(value)
            continue
        if key_norm in failed_keys:
            failed_values.append(value)
            continue
        if key_norm in not_attended_keys:
            not_attended_values.append(value)
            continue

        subject_rows.append((str(raw_key).strip(), value))

    metric_values = {
        "attendance": attendance_values,
        "failed_subjects": failed_values,
        "not_attended": not_attended_values,
        "gpa": gpa_values,
    }
    metric_labels = {
        "attendance": "Attendance",
        "failed_subjects": "Failed Subjects",
        "not_attended": "Not Attended",
        "gpa": "GPA",
    }

    used_subject_idx = set()
    ordered_lines = []

    def _pop_metric(metric_key):
        values = metric_values.get(metric_key) or []
        if not values:
            return None
        return values.pop(0)

    def _match_subject(raw_key, normalized_key):
        raw_key = str(raw_key or "").strip().lower()
        normalized_key = str(normalized_key or "").strip().lower()
        for idx, (label, value) in enumerate(subject_rows):
            if idx in used_subject_idx:
                continue
            label_norm = _normalize_metric_key(label)
            if raw_key and label.lower() == raw_key:
                used_subject_idx.add(idx)
                return label, value
            if normalized_key and label_norm == normalized_key:
                used_subject_idx.add(idx)
                return label, value
        return None

    if isinstance(ordered_fields, list):
        for item in ordered_fields:
            if not isinstance(item, dict):
                continue
            item_type = str(item.get("type") or "").strip().lower()
            if item_type == "metric":
                metric_key = str(item.get("key") or "").strip().lower()
                value = _pop_metric(metric_key)
                if value is not None:
                    ordered_lines.append(f"{metric_labels.get(metric_key, metric_key)} : {value}")
            elif item_type == "subject":
                subject = _match_subject(item.get("raw_key"), item.get("normalized_key"))
                if subject:
                    ordered_lines.append(f"{subject[0]} : {subject[1]}")

    if isinstance(ordered_fields, list):
        for idx, (subject, value) in enumerate(subject_rows):
            if idx not in used_subject_idx:
                ordered_lines.append(f"{subject} : {value}")

        for key in ("attendance", "failed_subjects", "not_attended", "gpa"):
            values = metric_values.get(key) or []
            for value in values:
                ordered_lines.append(f"{metric_labels[key]} : {value}")
    else:
        for value in attendance_values:
            ordered_lines.append(f"{metric_labels['attendance']} : {value}")
        for subject, value in subject_rows:
            ordered_lines.append(f"{subject} : {value}")
        for key in ("failed_subjects", "not_attended", "gpa"):
            for value in metric_values.get(key) or []:
                ordered_lines.append(f"{metric_labels[key]} : {value}")

    lines = []
    lines.extend(ordered_lines)

    return "\n".join(lines)


def _build_parent_message(test_name, reg_no, student_name, marks):
    marks_table = _build_parent_subjects_table(marks)
    return (
        f"Dear Parent , The Following is the {test_name} Marks Secured in each Course by your son/daughter\n\n"
        f"REGISTER NUMBER :  {reg_no}\n"
        f"NAME : {student_name}\n\n"
        f"{marks_table}\n\n"
        f"Regards\n"
        f"PRINCIPAL\n"
        f"RMKCET"
    )


# ============================= PAGE ROUTES =================================

@app.route("/")
def index():
    if "user_email" in session:
        if _is_admin_portal_user(session.get("role")):
            return redirect(url_for(_panel_endpoint_for_role(session.get("role"))))
        return redirect(url_for("counselor_page"))
    return redirect(url_for("login"))


def _parse_iso_datetime(value):
    try:
        return datetime.fromisoformat(str(value))
    except Exception:
        return None


def _is_pending_otp_valid(payload):
    expires_at = _parse_iso_datetime((payload or {}).get("expires_at"))
    return bool(expires_at and datetime.now() <= expires_at)


def _otp_hash(value):
    return hashlib.sha256(str(value or "").strip().encode("utf-8")).hexdigest()


def _complete_login_for_user(user):
    sid = str(uuid.uuid4())
    ok, msg = db.register_session(sid, user["email"], request.remote_addr, request.user_agent.string)
    if not ok:
        flash(msg, "error")
        return render_template("login.html")

    session["user_email"] = user["email"]
    session["user_name"] = user["name"]
    session["role"] = user["role"]
    session["session_id"] = sid
    session["department"] = user.get("department", "")

    show_tutorial_welcome = (not user.get("last_login")) and _is_tutorial_enabled_for_role(user.get("role"))
    redirect_params = {"tutorial_welcome": "1"} if show_tutorial_welcome else {}

    if _is_admin_portal_user(user["role"]):
        return redirect(url_for(_panel_endpoint_for_role(user["role"]), **redirect_params))
    return redirect(url_for("counselor_page", **redirect_params))


def _render_login_otp_challenge(extra_context=None):
    pending = session.get("pending_login_otp") or {}
    ctx = {
        "otp_challenge": True,
        "otp_email_masked": _mask_email(pending.get("email", "")),
    }
    if isinstance(extra_context, dict):
        ctx.update(extra_context)
    return render_template("login.html", **ctx)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        identifier = request.form.get("identifier", "").strip()  # Can be email or name
        password = request.form.get("password", "")
        force_logout = request.form.get("force_logout") == "true"

        user = db.authenticate_user(identifier, password)
        if not user:
            flash("Invalid email/name or password.", "error")
            return render_template("login.html", login_identifier=identifier)

        email = user["email"]
        allowed, msg = db.check_user_access(email)
        if not allowed:
            flash(msg, "error")
            return render_template("login.html", login_identifier=identifier)

        # Check for existing active session on another device
        if not force_logout:
            existing_session = db.get_user_active_session(email)
            if existing_session:
                # Return with session conflict info
                return render_template("login.html", 
                    session_conflict=True,
                    existing_session=existing_session,
                    stored_identifier=identifier,
                    stored_password=password)

        # Force logout existing sessions if requested
        if force_logout:
            db.force_logout_by_email(email, "new_device_login")

        if _is_login_otp_required_for_user(user.get("role"), email):
            otp_code = generate_otp(6)
            if not _send_otp_email(email, otp_code, "Login"):
                flash("OTP delivery failed. Verify SMTP settings and try again.", "error")
                return render_template("login.html", login_identifier=identifier)

            session["pending_login_otp"] = {
                "email": email,
                "otp_hash": _otp_hash(otp_code),
                "expires_at": (datetime.now() + timedelta(seconds=OTP_EXPIRY_SECONDS)).isoformat(),
                "requested_at": datetime.now().isoformat(),
            }
            session.modified = True
            flash(f"Login OTP sent to {_mask_email(email)}.", "info")
            return _render_login_otp_challenge()

        return _complete_login_for_user(user)

    pending_login = session.get("pending_login_otp")
    if pending_login:
        if _is_pending_otp_valid(pending_login):
            return _render_login_otp_challenge()
        session.pop("pending_login_otp", None)
        flash("Previous login OTP has expired. Please sign in again.", "warning")

    return render_template("login.html")


@app.route("/login/otp", methods=["POST"])
def verify_login_otp():
    pending = session.get("pending_login_otp") or {}
    code = str(request.form.get("otp_code") or "").strip()
    if not pending:
        flash("No pending OTP challenge. Please login again.", "error")
        return redirect(url_for("login"))

    if not _is_pending_otp_valid(pending):
        session.pop("pending_login_otp", None)
        flash("OTP expired. Please login again.", "error")
        return redirect(url_for("login"))

    if not code or _otp_hash(code) != str(pending.get("otp_hash") or ""):
        flash("Invalid OTP. Please try again.", "error")
        return _render_login_otp_challenge()

    email = str(pending.get("email") or "").strip()
    user = db.get_user(email)
    if not user:
        session.pop("pending_login_otp", None)
        flash("User not found. Please login again.", "error")
        return redirect(url_for("login"))

    allowed, msg = db.check_user_access(email)
    if not allowed:
        session.pop("pending_login_otp", None)
        flash(msg, "error")
        return redirect(url_for("login"))

    session.pop("pending_login_otp", None)
    return _complete_login_for_user(user)


@app.route("/login/otp/resend", methods=["POST"])
def resend_login_otp():
    pending = session.get("pending_login_otp") or {}
    if not pending:
        flash("No pending OTP challenge. Please login again.", "error")
        return redirect(url_for("login"))

    last_requested = _parse_iso_datetime(pending.get("requested_at"))
    if last_requested and (datetime.now() - last_requested).total_seconds() < 30:
        flash("Please wait 30 seconds before requesting another OTP.", "warning")
        return _render_login_otp_challenge()

    email = str(pending.get("email") or "").strip()
    user = db.get_user(email)
    if not user:
        session.pop("pending_login_otp", None)
        flash("User not found. Please login again.", "error")
        return redirect(url_for("login"))

    otp_code = generate_otp(6)
    if not _send_otp_email(email, otp_code, "Login"):
        flash("Unable to resend OTP. Please try login again.", "error")
        return _render_login_otp_challenge()

    pending["otp_hash"] = _otp_hash(otp_code)
    pending["expires_at"] = (datetime.now() + timedelta(seconds=OTP_EXPIRY_SECONDS)).isoformat()
    pending["requested_at"] = datetime.now().isoformat()
    session["pending_login_otp"] = pending
    session.modified = True
    flash(f"A new OTP was sent to {_mask_email(email)}.", "info")
    return _render_login_otp_challenge()


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    pending = session.get("pending_password_reset") or {}

    if request.method == "POST":
        action = str(request.form.get("action") or "request").strip().lower()

        if action == "request":
            identifier = str(request.form.get("identifier") or "").strip()
            user = db.get_user_by_identifier(identifier)
            if not user:
                flash("No account found for the provided email/name.", "error")
                return render_template("forgot_password.html", stage="request")

            email = str(user.get("email") or "").strip()
            if not email:
                flash("Account email is not configured. Contact system admin.", "error")
                return render_template("forgot_password.html", stage="request")

            otp_required = _is_password_reset_otp_required_for_user(user.get("role"), email)
            reset_payload = {
                "email": email,
                "verified": False,
                "expires_at": (datetime.now() + timedelta(minutes=15)).isoformat(),
            }

            if otp_required:
                otp_code = generate_otp(6)
                if not _send_otp_email(email, otp_code, "Password Reset"):
                    flash("Unable to send OTP. Please try again later.", "error")
                    return render_template("forgot_password.html", stage="request")
                reset_payload["otp_hash"] = _otp_hash(otp_code)
                reset_payload["otp_expires_at"] = (datetime.now() + timedelta(seconds=OTP_EXPIRY_SECONDS)).isoformat()
                flash(f"Password reset OTP sent to {_mask_email(email)}.", "info")
            else:
                reset_payload["verified"] = True

            session["pending_password_reset"] = reset_payload
            session.modified = True
            return redirect(url_for("forgot_password"))

        if action == "verify":
            if not pending:
                flash("Start password reset again.", "error")
                return redirect(url_for("forgot_password"))

            otp_code = str(request.form.get("otp_code") or "").strip()
            otp_expires = _parse_iso_datetime(pending.get("otp_expires_at"))
            if not otp_expires or datetime.now() > otp_expires:
                session.pop("pending_password_reset", None)
                flash("OTP expired. Request a new password reset.", "error")
                return redirect(url_for("forgot_password"))

            if not otp_code or _otp_hash(otp_code) != str(pending.get("otp_hash") or ""):
                flash("Invalid OTP.", "error")
                return render_template("forgot_password.html", stage="verify", masked_email=_mask_email(pending.get("email")))

            pending["verified"] = True
            pending.pop("otp_hash", None)
            pending.pop("otp_expires_at", None)
            session["pending_password_reset"] = pending
            session.modified = True
            flash("OTP verified. Set your new password.", "success")
            return redirect(url_for("forgot_password"))

        if action == "reset":
            if not pending:
                flash("Start password reset again.", "error")
                return redirect(url_for("forgot_password"))

            expires_at = _parse_iso_datetime(pending.get("expires_at"))
            if not expires_at or datetime.now() > expires_at:
                session.pop("pending_password_reset", None)
                flash("Reset session expired. Start again.", "error")
                return redirect(url_for("forgot_password"))

            if not bool(pending.get("verified")):
                flash("Verify OTP before resetting password.", "error")
                return redirect(url_for("forgot_password"))

            new_password = str(request.form.get("new_password") or "").strip()
            confirm_password = str(request.form.get("confirm_password") or "").strip()
            if len(new_password) < 6:
                flash("Password must be at least 6 characters.", "error")
                return render_template("forgot_password.html", stage="reset", masked_email=_mask_email(pending.get("email")))
            if new_password != confirm_password:
                flash("Password and confirm password do not match.", "error")
                return render_template("forgot_password.html", stage="reset", masked_email=_mask_email(pending.get("email")))

            target_email = str(pending.get("email") or "").strip()
            user = db.get_user(target_email)
            if not user:
                session.pop("pending_password_reset", None)
                flash("User not found.", "error")
                return redirect(url_for("forgot_password"))

            db.update_user(target_email, password=new_password)
            db.force_logout_user(target_email, "self_password_reset")
            session.pop("pending_password_reset", None)
            flash("Password reset successful. Please login with the new password.", "success")
            return redirect(url_for("login"))

    if pending:
        if bool(pending.get("verified")):
            return render_template("forgot_password.html", stage="reset", masked_email=_mask_email(pending.get("email")))
        if pending.get("otp_hash"):
            return render_template("forgot_password.html", stage="verify", masked_email=_mask_email(pending.get("email")))

    return render_template("forgot_password.html", stage="request")


@app.route("/logout")
def logout():
    sid = session.get("session_id")
    if sid:
        db.end_session(sid)
    session.clear()
    return redirect(url_for("login"))


# ======================== ADMIN PAGE =======================================

@app.route("/admin")
@login_required
@admin_required
def admin():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if not _is_system_admin(actor_role):
        return redirect(url_for(_panel_endpoint_for_role(actor_role), **request.args.to_dict(flat=True)))

    return _render_admin_panel(actor_email, actor_role)


@app.route("/principal")
@login_required
@admin_required
def principal_panel():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    preview_mode = _is_system_admin(actor_role) and session.get("ui_preview_role") == "principal"
    if not _is_principal(actor_role) and not preview_mode:
        return redirect(url_for(_panel_endpoint_for_role(actor_role)))
    render_role = "principal" if preview_mode else actor_role
    return _render_admin_panel(actor_email, render_role, preview_mode=preview_mode)


@app.route("/hod")
@login_required
@admin_required
def hod_panel():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    preview_mode = _is_system_admin(actor_role) and session.get("ui_preview_role") == "hod"
    if not _is_hod(actor_role) and not preview_mode:
        return redirect(url_for(_panel_endpoint_for_role(actor_role)))
    forced_scope_pairs = _build_example_scope_pairs() if preview_mode else None
    return _render_admin_panel(actor_email, "hod" if preview_mode else actor_role, preview_mode=preview_mode, forced_scope_pairs=forced_scope_pairs)


@app.route("/deo")
@login_required
@admin_required
def deo_panel():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    preview_mode = _is_system_admin(actor_role) and session.get("ui_preview_role") == "deo"
    if not _is_deo(actor_role) and not preview_mode:
        return redirect(url_for(_panel_endpoint_for_role(actor_role)))
    forced_scope_pairs = _build_example_scope_pairs() if preview_mode else None
    return _render_admin_panel(actor_email, "deo" if preview_mode else actor_role, preview_mode=preview_mode, forced_scope_pairs=forced_scope_pairs)


@app.route("/admin/test-mode/<preview_role>")
@login_required
@admin_required
def admin_test_mode_preview(preview_role):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if not _is_system_admin(actor_role):
        flash("Only system admin can use test mode previews.", "error")
        return _redirect_admin_back("config")

    role_norm = str(preview_role or "").strip().lower()
    if role_norm in {"principal", "hod", "deo", "counselor"}:
        session["ui_preview_role"] = role_norm

    if role_norm == "principal":
        return redirect(url_for("principal_panel", tab=request.args.get("tab") or "dashboard"))
    if role_norm == "hod":
        return redirect(url_for("hod_panel", tab=request.args.get("tab") or "dashboard"))
    if role_norm == "deo":
        return redirect(url_for("deo_panel", tab=request.args.get("tab") or "reports"))
    if role_norm == "counselor":
        return redirect(url_for("counselor_page", tab=request.args.get("tab") or "recent-tests"))

    flash("Unsupported preview role.", "error")
    return _redirect_admin_back("config")


@app.route("/admin/test-mode/exit")
@login_required
@admin_required
def admin_test_mode_exit():
    if _is_system_admin(session.get("role")):
        session.pop("ui_preview_role", None)
        flash("Exited test mode preview.", "success")
    return redirect(url_for("admin", tab="config"))


def _render_admin_panel(actor_email, actor_role, preview_mode=False, forced_scope_pairs=None):
    requested_tab = (request.args.get("tab") or "").strip()
    current_tab = requested_tab or _default_tab_for_role(actor_role)

    allowed_tabs = _allowed_tabs_for_role(actor_role)
    if current_tab not in allowed_tabs:
        fallback = "dashboard" if (_is_principal(actor_role) or _is_hod(actor_role)) else "reports"
        if requested_tab:
            flash("Access denied for this section.", "warning")
        return redirect(url_for(_panel_endpoint_for_role(actor_role), tab=fallback))

    if forced_scope_pairs is not None:
        allowed_scopes = {(str(dep).upper(), int(yr)) for dep, yr in forced_scope_pairs}
    else:
        allowed_scopes = _get_actor_scope_pairs(actor_email, actor_role)

    users = db.get_scoped_users_for_admin(actor_email, actor_role)
    departments = db.get_departments_for_admin(actor_email, actor_role, active_only=False)
    if forced_scope_pairs is not None:
        all_departments = db.get_departments(active_only=False)
        allowed_department_codes = {dep for dep, _ in forced_scope_pairs}
        departments = [d for d in all_departments if str(d.get("code") or "").strip().upper() in allowed_department_codes]
    active_sessions = db.get_active_sessions() if _is_system_admin(actor_role) else []
    full_activity = db.get_counselor_activity_summary()
    activity = _filter_activity_for_actor(full_activity, actor_email, actor_role)
    format_settings = db.get_format_settings()

    msg_day = (request.args.get("msg_day") or "").strip()
    msg_q = (request.args.get("msg_q") or "").strip()
    msg_year = (request.args.get("msg_year") or "").strip()
    msg_month = (request.args.get("msg_month") or "").strip()
    msg_day_num = (request.args.get("msg_day_num") or "").strip()
    allowed_counselors = _get_allowed_counselor_emails_for_actor(
        actor_email,
        actor_role,
        forced_scope_pairs=forced_scope_pairs,
    )

    messages = db.get_message_history_filtered(
        day=msg_day or None,
        counselor_query=msg_q or None,
        limit=1500,
        filter_year=msg_year or None,
        filter_month=msg_month or None,
        filter_day=msg_day_num or None,
        allowed_counselors=allowed_counselors,
    )
    message_days = db.get_message_days(counselor_query=msg_q or None)
    grouped_map = {}
    for m in messages:
        day_key = str((m.get("sent_at") or "")[:10] or "Unknown")
        grouped_map.setdefault(day_key, []).append(m)
    message_groups = [{"day": day, "messages": rows, "total": len(rows)} for day, rows in grouped_map.items()]
    msg_stats = db.get_message_stats(actor_email) if (_is_hod(actor_role) or _is_deo(actor_role)) else db.get_message_stats()
    counselor_suggestions = db.get_message_counselor_suggestions(
        query=msg_q,
        allowed_counselors=allowed_counselors,
    )
    
    # App configuration
    app_config = db.get_app_config()
    
    # Session monitoring statistics
    session_stats = db.get_session_statistics() if _is_system_admin(actor_role) else {
        "active_sessions": 0, "today_sessions": 0, "avg_duration_minutes": 0, "forced_logouts": 0,
        "peak_concurrent": 0, "desktop_sessions": 0, "mobile_sessions": 0, "logout_reasons": {}
    }
    session_history = db.get_session_history(limit=100) if _is_system_admin(actor_role) else []
    session_log_ok = True
    session_log_message = ""
    if _is_system_admin(actor_role):
        log_path = os.path.join(DATA_DIR, "session_history.log")
        session_log_ok, session_log_message = db.rotate_session_log_daily(log_path)

    selected_department = (request.args.get("dept") or "").strip().upper()
    selected_year_level = request.args.get("year_level", type=int) or 1
    department_tests = db.get_all_unique_tests(
        filter_dept=selected_department or None,
        filter_year_level=selected_year_level,
        allowed_scopes=allowed_scopes,
    )

    is_scoped_admin_role = _is_hod(actor_role) or _is_deo(actor_role)
    if forced_scope_pairs is not None and actor_role in {"hod", "deo"}:
        scoped_rows = [{"department": dep, "year_level": yr} for dep, yr in forced_scope_pairs]
    else:
        scoped_rows = db.get_chief_admin_scopes(actor_email) if is_scoped_admin_role else []

    scoped_department_codes, scoped_years_by_department = _scope_pairs_to_nav_map(
        [(str(s.get("department") or "").strip().upper(), int(s.get("year_level") or 1)) for s in scoped_rows]
    )

    report_department = (request.args.get("report_dept") or selected_department or "").strip().upper()
    if is_scoped_admin_role and report_department and report_department not in scoped_department_codes:
        report_department = ""

    report_available_years = (
        scoped_years_by_department.get(report_department, [])
        if (is_scoped_admin_role and report_department)
        else [1, 2, 3, 4]
    )
    report_year_level = request.args.get("report_year", type=int)
    if report_year_level not in report_available_years:
        report_year_level = None

    report_tests = []
    if report_department and report_year_level in (1, 2, 3, 4):
        report_tests = db.get_all_unique_tests(
            filter_dept=report_department,
            filter_year_level=report_year_level,
            allowed_scopes=allowed_scopes,
        )
    recent_report_tests = db.get_all_unique_tests(allowed_scopes=allowed_scopes)[:6]

    # Counselor activity filter model
    act_department = (request.args.get("act_dept") or report_department or "").strip().upper()
    if is_scoped_admin_role and act_department and act_department not in scoped_department_codes:
        act_department = ""

    activity_available_years = (
        scoped_years_by_department.get(act_department, [])
        if (is_scoped_admin_role and act_department)
        else [1, 2, 3, 4]
    )
    act_year_level = request.args.get("act_year", type=int)
    if act_year_level not in activity_available_years:
        act_year_level = None
    act_semester = (request.args.get("act_sem") or "").strip()
    act_test_name = _normalize_allowed_test_name(request.args.get("act_test") or "")
    act_search = (request.args.get("act_q") or "").strip()
    act_sort = (request.args.get("act_sort") or "pending_first").strip()
    activity_selection_ready = False
    activity_test_result = {"rows": [], "stats": {"total_counselors": 0, "complete": 0, "pending": 0, "avg_completion": 0}}

    activity_test_status = {
        "1": {"IAT 1": False, "IAT 2": False, "MODEL EXAM": False},
        "2": {"IAT 1": False, "IAT 2": False, "MODEL EXAM": False},
    }
    activity_has_scope = bool(act_department and act_year_level in activity_available_years)
    if activity_has_scope:
        uploaded = db.get_all_unique_tests(
            filter_dept=act_department,
            filter_year_level=act_year_level,
            allowed_scopes=allowed_scopes,
        )
        for t in uploaded:
            sem = str(t.get("semester") or "").strip()
            test_name = _normalize_allowed_test_name(t.get("test_name") or "")
            if sem in activity_test_status and test_name in activity_test_status[sem]:
                activity_test_status[sem][test_name] = True

    activity_show_department = not act_department
    activity_show_year = bool(act_department) and (act_year_level not in activity_available_years)
    activity_show_semester = activity_has_scope and not activity_selection_ready

    if act_department and act_year_level in (1, 2, 3, 4) and act_semester in {"1", "2"} and act_test_name:
        activity_selection_ready = True
        activity_test_result = db.get_counselor_activity_for_test(
            department=act_department,
            year_level=act_year_level,
            semester=act_semester,
            test_name=act_test_name,
            search_query=act_search,
            sort_mode=act_sort,
        )

    # Backup inventory for database tab
    backup_dir = os.path.join(DATA_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    backup_files = []
    for name in sorted(os.listdir(backup_dir), reverse=True):
        path = os.path.join(backup_dir, name)
        if not os.path.isfile(path):
            continue
        stat = os.stat(path)
        backup_files.append({
            "name": name,
            "size_kb": int(stat.st_size / 1024),
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
        })
    chief_scopes = scoped_rows if (_is_hod(actor_role) or _is_deo(actor_role)) else []
    chief_scope_keys = [
        f"{str(s.get('department') or '').upper()}::{int(s.get('year_level') or 1)}"
        for s in chief_scopes
    ]
    chief_scopes_by_email = {}
    for u in users:
        if u.get("role") not in {"chief_admin", "hod", "deo"}:
            continue
        scopes_for_user = db.get_chief_admin_scopes(u.get("email"))
        chief_scopes_by_email[u.get("email")] = [
            f"{str(s.get('department') or '').upper()}::{int(s.get('year_level') or 1)}"
            for s in scopes_for_user
        ]

    counselors = [u for u in users if u["role"] == "counselor"]
    dashboard_data = _build_admin_dashboard_data(actor_email, actor_role, activity, departments)
    leaderboard = sorted(
        [a for a in activity if int(a.get("student_count") or 0) > 0],
        key=lambda r: (
            -float((int(r.get("unique_students_messaged") or 0) / max(1, int(r.get("student_count") or 1))) * 100),
            int(r.get("best_completion_seconds") or 10**9),
            -int(r.get("total_messages") or 0),
        ),
    )[:10]
    students_map = {c["email"]: db.get_students(c["email"]) for c in counselors}
    return render_template(
        "admin.html",
        users=users,
        departments=departments,
        sessions=active_sessions,
        activity=activity,
        format_settings=format_settings,
        messages=messages,
        message_groups=message_groups,
        message_days=message_days,
        selected_message_day=msg_day,
        message_query=msg_q,
        selected_message_year=msg_year,
        selected_message_month=msg_month,
        selected_message_day_num=msg_day_num,
        counselor_suggestions=counselor_suggestions,
        msg_stats=msg_stats,
        counselor_count=len(counselors),
        active_counselor_count=sum(1 for c in counselors if c["is_active"]),
        session_count=len(active_sessions),
        department_tests=department_tests,
        selected_department=selected_department,
        selected_year_level=selected_year_level,
        report_department=report_department,
        report_year_level=report_year_level,
        report_tests=report_tests,
        report_available_years=report_available_years,
        recent_report_tests=recent_report_tests,
        chief_scopes=chief_scopes,
        chief_scope_keys=chief_scope_keys,
        chief_scopes_by_email=chief_scopes_by_email,
        is_system_admin=_is_system_admin(actor_role),
        is_chief_admin=_is_hod(actor_role),
        is_deo=_is_deo(actor_role),
        is_principal=_is_principal(actor_role),
        app_config=app_config,
        session_stats=session_stats,
        session_history=session_history,
        session_log_ok=session_log_ok,
        session_log_message=session_log_message,
        students_map=students_map,
        dashboard_data=dashboard_data,
        leaderboard=leaderboard,
        activity_selection_ready=activity_selection_ready,
        activity_test_result=activity_test_result,
        act_department=act_department,
        act_year_level=act_year_level,
        activity_available_years=activity_available_years,
        act_semester=act_semester,
        act_test_name=act_test_name,
        act_search=act_search,
        act_sort=act_sort,
        backup_files=backup_files,
        default_tab=_default_tab_for_role(actor_role),
        panel_endpoint=_panel_endpoint_for_role(actor_role),
        activity_test_status=activity_test_status,
        activity_has_scope=activity_has_scope,
        activity_show_department=activity_show_department,
        activity_show_year=activity_show_year,
        activity_show_semester=activity_show_semester,
        current_role_override=(actor_role if preview_mode else ""),
        preview_mode=preview_mode,
    )


def _perform_database_backup(batch_name, overwrite=False):
    backup_dir = os.path.join(DATA_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    safe_batch = re.sub(r"[^0-9A-Za-z_-]", "_", str(batch_name or "batch"))
    target = os.path.join(backup_dir, f"rmkcet_shine_{safe_batch}.db")
    if os.path.exists(target) and not overwrite:
        raise ValueError("Backup for this batch already exists. Enable overwrite to replace it.")
    shutil.copy2(DATABASE_FILE, target)
    return target


def _clear_exam_database_only():
    conn = sqlite3.connect(DATABASE_FILE)
    try:
        conn.execute("PRAGMA foreign_keys = OFF")
        conn.execute("DELETE FROM counselor_mark_overrides")
        conn.execute("DELETE FROM sent_messages")
        conn.execute("DELETE FROM student_marks")
        conn.execute("DELETE FROM test_metadata")
        conn.execute("DELETE FROM tests")
        conn.execute("DELETE FROM semesters")
        conn.execute("DELETE FROM batches")
        conn.commit()
    finally:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.close()


def _validate_current_user_password(password):
    user = db.get_user(session.get("user_email") or "") or {}
    return db.verify_password(str(password or ""), user.get("password_hash") or "")


def _latest_backup_is_recent_enough():
    backup_dir = os.path.join(DATA_DIR, "backups")
    if not os.path.isdir(backup_dir):
        return False
    files = [os.path.join(backup_dir, n) for n in os.listdir(backup_dir) if os.path.isfile(os.path.join(backup_dir, n))]
    if not files:
        return False
    latest = max(files, key=os.path.getmtime)
    return os.path.getmtime(latest) >= (os.path.getmtime(DATABASE_FILE) - 2)


@app.route("/api/database/backup", methods=["POST"])
@login_required
@admin_required
def api_database_backup_create():
    role = session.get("role")
    if not (_is_system_admin(role) or _is_principal(role)):
        flash("Only system admin or principal can perform database maintenance.", "error")
        return _redirect_admin_back("reports")

    batch_name = (request.form.get("batch_name") or "").strip() or f"{datetime.now().year}-{datetime.now().year + 1}"
    overwrite = request.form.get("overwrite") == "on"
    try:
        backup_path = _perform_database_backup(batch_name, overwrite=overwrite)
        flash(f"Backup created at {backup_path}", "success")
    except Exception as e:
        flash(f"Database maintenance failed: {e}", "error")
    return _redirect_admin_back("database")


@app.route("/api/database/clear", methods=["POST"])
@login_required
@admin_required
def api_database_clear_exam_data():
    role = session.get("role")
    if not (_is_system_admin(role) or _is_principal(role)):
        flash("Only system admin or principal can perform database maintenance.", "error")
        return _redirect_admin_back("reports")

    password = request.form.get("admin_password") or request.form.get("password", "")
    if not _validate_current_user_password(password):
        flash("Password verification failed.", "error")
        return _redirect_admin_back("database")

    if not _latest_backup_is_recent_enough():
        flash("Clear blocked: create a latest backup before clearing exam data.", "error")
        return _redirect_admin_back("database")

    try:
        db.logout_all_users()
        _clear_exam_database_only()
        session.clear()
        flash("Exam database cleared successfully.", "success")
        return redirect(url_for("login"))
    except Exception as e:
        flash(f"Clear failed: {e}", "error")
        return _redirect_admin_back("database")


@app.route("/api/database/restore", methods=["POST"])
@login_required
@admin_required
def api_database_restore_backup():
    role = session.get("role")
    if not (_is_system_admin(role) or _is_principal(role)):
        flash("Only system admin or principal can perform database maintenance.", "error")
        return _redirect_admin_back("reports")

    password = request.form.get("admin_password") or request.form.get("password", "")
    backup_name = (request.form.get("backup_name") or "").strip()
    if not _validate_current_user_password(password):
        flash("Password verification failed.", "error")
        return _redirect_admin_back("database")
    if not backup_name:
        flash("Select a backup to restore.", "error")
        return _redirect_admin_back("database")

    backup_path = os.path.join(DATA_DIR, "backups", backup_name)
    if not os.path.isfile(backup_path):
        flash("Backup file not found.", "error")
        return _redirect_admin_back("database")

    try:
        db.logout_all_users()
        shutil.copy2(backup_path, DATABASE_FILE)
        session.clear()
        flash("Database restored from backup.", "success")
        return redirect(url_for("login"))
    except Exception as e:
        flash(f"Restore failed: {e}", "error")
        return _redirect_admin_back("database")


@app.route("/api/database/delete-backup", methods=["POST"])
@login_required
@admin_required
def api_database_delete_backup():
    role = session.get("role")
    if not (_is_system_admin(role) or _is_principal(role)):
        flash("Only system admin or principal can perform database maintenance.", "error")
        return _redirect_admin_back("reports")

    password = request.form.get("admin_password") or request.form.get("password", "")
    backup_name = (request.form.get("backup_name") or "").strip()
    if not _validate_current_user_password(password):
        flash("Password verification failed.", "error")
        return _redirect_admin_back("database")

    backup_path = os.path.join(DATA_DIR, "backups", backup_name)
    if not os.path.isfile(backup_path):
        flash("Backup file not found.", "error")
        return _redirect_admin_back("database")

    try:
        os.remove(backup_path)
        flash("Backup deleted.", "success")
    except Exception as e:
        flash(f"Delete failed: {e}", "error")
    return _redirect_admin_back("database")


@app.route("/api/messages/delete/<int:message_id>", methods=["POST"])
@login_required
@admin_required
def api_delete_message(message_id):
    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    deleted = db.delete_message_by_id(message_id)
    if deleted:
        flash("Message entry deleted.", "success")
    else:
        flash("Message entry not found.", "warning")
    return _redirect_admin_back("messages", msg_day=msg_day, msg_q=msg_q, msg_year=msg_year, msg_month=msg_month, msg_day_num=msg_day_num)


@app.route("/api/messages/delete-bulk", methods=["POST"])
@login_required
@admin_required
def api_delete_messages_bulk():
    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    ids = request.form.getlist("message_ids")
    if not ids:
        flash("Select at least one message to delete.", "warning")
        return _redirect_admin_back("messages", msg_day=msg_day, msg_q=msg_q, msg_year=msg_year, msg_month=msg_month, msg_day_num=msg_day_num)

    deleted = db.delete_messages_by_ids(ids)
    if deleted:
        flash(f"Deleted {deleted} message entr{'y' if deleted == 1 else 'ies'}.", "success")
    else:
        flash("No messages were deleted.", "warning")
    return _redirect_admin_back("messages", msg_day=msg_day, msg_q=msg_q, msg_year=msg_year, msg_month=msg_month, msg_day_num=msg_day_num)


@app.route("/api/messages/export/csv")
@login_required
@admin_required
def api_export_messages_csv():
    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    allowed_counselors = _get_allowed_counselor_emails_for_actor(session.get("user_email"), session.get("role"))
    data = db.get_message_history_filtered(
        day=msg_day or None,
        counselor_query=msg_q or None,
        limit=None,
        filter_year=msg_year or None,
        filter_month=msg_month or None,
        filter_day=msg_day_num or None,
        allowed_counselors=allowed_counselors,
    )

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Date", "Time", "Counselor", "Counselor Email", "Student", "Reg No", "Format", "Status"])
    for m in data:
        sent_at = str(m.get("sent_at") or "")
        date_part = sent_at[:10]
        time_part = sent_at[11:19] if len(sent_at) >= 19 else ""
        w.writerow([
            date_part,
            time_part,
            m.get("counselor_name") or m.get("counselor_email") or "",
            m.get("counselor_email") or "",
            m.get("student_name") or "",
            m.get("reg_no") or "",
            m.get("format") or "",
            m.get("status") or "",
        ])

    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={_message_export_filename('csv')}"},
    )


@app.route("/api/messages/export/excel")
@login_required
@admin_required
def api_export_messages_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    allowed_counselors = _get_allowed_counselor_emails_for_actor(session.get("user_email"), session.get("role"))
    data = db.get_message_history_filtered(
        day=msg_day or None,
        counselor_query=msg_q or None,
        limit=None,
        filter_year=msg_year or None,
        filter_month=msg_month or None,
        filter_day=msg_day_num or None,
        allowed_counselors=allowed_counselors,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Message Activity"

    headers = ["Date", "Time", "Counselor", "Counselor Email", "Student", "Reg No", "Format", "Status"]
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, m in enumerate(data, 2):
        sent_at = str(m.get("sent_at") or "")
        ws.cell(row=ri, column=1, value=sent_at[:10])
        ws.cell(row=ri, column=2, value=sent_at[11:19] if len(sent_at) >= 19 else "")
        ws.cell(row=ri, column=3, value=m.get("counselor_name") or m.get("counselor_email") or "")
        ws.cell(row=ri, column=4, value=m.get("counselor_email") or "")
        ws.cell(row=ri, column=5, value=m.get("student_name") or "")
        ws.cell(row=ri, column=6, value=m.get("reg_no") or "")
        ws.cell(row=ri, column=7, value=m.get("format") or "")
        ws.cell(row=ri, column=8, value=m.get("status") or "")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 36)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=_message_export_filename("xlsx"),
    )


@app.route("/api/messages/export/pdf")
@login_required
@admin_required
def api_export_messages_pdf():
    msg_day, msg_q, msg_year, msg_month, msg_day_num = _get_message_filters_from_request()
    allowed_counselors = _get_allowed_counselor_emails_for_actor(session.get("user_email"), session.get("role"))
    data = db.get_message_history_filtered(
        day=msg_day or None,
        counselor_query=msg_q or None,
        limit=None,
        filter_year=msg_year or None,
        filter_month=msg_month or None,
        filter_day=msg_day_num or None,
        allowed_counselors=allowed_counselors,
    )

    pdf = FPDF("L")
    pdf.add_page()
    pdf.set_font("Arial", "B", 15)
    pdf.cell(0, 10, "RMKCET Parent Connect - Message Activity", 0, 1, "C")
    pdf.set_font("Arial", "", 9)
    subtitle = f"Day: {msg_day or 'All'}   Counselor Search: {msg_q or 'All'}   Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    pdf.cell(0, 8, subtitle[:150], 0, 1, "C")
    pdf.ln(2)

    widths = [24, 18, 44, 52, 36, 24, 20, 20]
    heads = ["Date", "Time", "Counselor", "Email", "Student", "Reg No", "Format", "Status"]

    pdf.set_font("Arial", "B", 8)
    pdf.set_fill_color(102, 126, 234)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(widths, heads):
        pdf.cell(w, 8, h, 1, 0, "C", True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 7)
    for m in data:
        sent_at = str(m.get("sent_at") or "")
        pdf.cell(widths[0], 7, sent_at[:10], 1)
        pdf.cell(widths[1], 7, (sent_at[11:19] if len(sent_at) >= 19 else ""), 1)
        pdf.cell(widths[2], 7, str(m.get("counselor_name") or m.get("counselor_email") or "")[:24], 1)
        pdf.cell(widths[3], 7, str(m.get("counselor_email") or "")[:30], 1)
        pdf.cell(widths[4], 7, str(m.get("student_name") or "")[:20], 1)
        pdf.cell(widths[5], 7, str(m.get("reg_no") or "")[:14], 1)
        pdf.cell(widths[6], 7, str(m.get("format") or "")[:10], 1, 0, "C")
        pdf.cell(widths[7], 7, str(m.get("status") or "")[:10], 1, 1, "C")

    buf = io.BytesIO()
    raw = pdf.output(dest="S")
    buf.write(raw if isinstance(raw, bytes) else raw.encode("latin-1"))
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=_message_export_filename("pdf"),
    )


# ======================== COUNSELOR PAGE ===================================

@app.route("/counselor")
@login_required
def counselor_page():
    email = session["user_email"]
    role = session.get("role")
    preview_mode = _is_system_admin(role) and session.get("ui_preview_role") == "counselor"

    if preview_mode:
        example_scope_pairs = _build_example_scope_pairs(limit=1)
        dep = example_scope_pairs[0][0] if example_scope_pairs else ""
        yr = example_scope_pairs[0][1] if example_scope_pairs else 1
        user = {
            "email": "preview.counselor@rmkcet.local",
            "name": "Example Counselor",
            "role": "counselor",
            "department": dep,
            "year_level": yr,
            "can_upload_students": 1,
        }
        tests = db.get_all_unique_tests(filter_dept=dep or None, filter_year_level=yr)[:20]
        recent_tests = tests[:2]
        msg_stats = {"total": 0, "week": 0}
        msg_history = []
        submissions = []
        selected_test_id = None
        selected_test_meta = None
        pending_students = []
        sent_reg_nos = set()
        return render_template(
            "counselor.html",
            user=user,
            is_blocked_department=False,
            students=[],
            assigned_students_count=0,
            tests=tests,
            recent_tests=recent_tests,
            msg_stats=msg_stats,
            msg_history=msg_history,
            submissions=submissions,
            selected_test_id=selected_test_id,
            selected_test_meta=selected_test_meta,
            pending_students=pending_students,
            sent_count=0,
            can_upload_students=True,
            report_tab=(request.args.get("tab") or "recent-tests"),
            counselor_dashboard_data={
                "student_activity": {"labels": [], "values": []},
                "test_histogram": {"labels": [], "values": [], "raw_values": [], "coverage": []},
            },
            current_role_override="counselor",
            preview_mode=True,
        )

    user = db.get_user(email)
    if not user:
        session.clear()
        return redirect(url_for("login"))

    is_blocked_department = not db.is_department_active(user.get("department"))
    students = db.get_students(email)
    tests = db.get_visible_tests_for_counselor(email)
    recent_tests = tests[:2]
    msg_stats = db.get_message_stats(email)
    msg_history = db.get_message_history(email, limit=50)
    
    # Get counselor's previous marksheet submissions
    submissions = db.get_counselor_submissions(email, limit=50)

    # Send-flow state (selected test and pending students)
    selected_test_id = request.args.get("test_id", type=int)
    valid_ids = {int(t.get("id")) for t in tests if t.get("id") is not None}
    if selected_test_id not in valid_ids:
        selected_test_id = int(tests[0]["id"]) if tests else None

    selected_test_meta = db.get_test_metadata(selected_test_id) if selected_test_id else None
    pending_students = db.get_pending_students_for_test(email, selected_test_id) if selected_test_id else students
    sent_reg_nos = db.get_sent_reg_nos_for_test(email, selected_test_id) if selected_test_id else set()
    counselor_dashboard_data = _build_counselor_dashboard_data(email)

    return render_template(
        "counselor.html",
        user=user,
        is_blocked_department=is_blocked_department,
        students=students,
        assigned_students_count=len(students),
        tests=tests,
        recent_tests=recent_tests,
        msg_stats=msg_stats,
        msg_history=msg_history,
        submissions=submissions,
        selected_test_id=selected_test_id,
        selected_test_meta=selected_test_meta,
        pending_students=pending_students,
        sent_count=(len(sent_reg_nos) if selected_test_id else 0),
        can_upload_students=bool(user.get("can_upload_students", 1)),
        report_tab=(request.args.get("tab") or "recent-tests"),
        counselor_dashboard_data=counselor_dashboard_data,
        current_role_override="",
        preview_mode=False,
    )


# ============================== API ROUTES ==================================

# ---------- Users -----------------------------------------------------------

def _resolve_department_from_branch(branch_value, department_codes):
    raw = str(branch_value or "").strip().upper()
    if not raw:
        return ""
    normalized = re.sub(r"\s+", "", raw)
    for code in department_codes:
        c = str(code or "").strip().upper()
        if not c:
            continue
        c_norm = re.sub(r"\s+", "", c)
        if c in raw or c_norm in normalized:
            return c

    token = re.split(r"[^A-Z0-9()]+", raw)[0].strip()
    return token


def _parse_bulk_counselor_excel(file_obj, department_codes):
    import pandas as pd

    xl = pd.ExcelFile(file_obj)
    parsed_rows = []
    for sheet_name in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet_name)
        if df is None or df.empty:
            continue

        # Normalize columns for resilient matching across template variations.
        col_map = {}
        for c in df.columns:
            key = re.sub(r"[^a-z0-9]", "", str(c or "").strip().lower())
            col_map[key] = c

        branch_col = next((col_map[k] for k in col_map if "branch" in k or "department" in k), None)
        name_col = next((col_map[k] for k in col_map if k in {"name", "counselorname"} or "name" in k), None)
        email_col = next((col_map[k] for k in col_map if "mailid" in k or k == "email" or "email" in k), None)
        password_col = next((col_map[k] for k in col_map if "password" in k), None)

        if not branch_col or not name_col or not email_col:
            continue

        for _, row in df.iterrows():
            name = str(row.get(name_col) or "").strip()
            email = str(row.get(email_col) or "").strip().lower()
            branch = str(row.get(branch_col) or "").strip()
            password = str(row.get(password_col) or "").strip() if password_col else ""

            if not name or not email or "@" not in email or not branch:
                continue

            dep = _resolve_department_from_branch(branch, department_codes)
            if not dep:
                continue

            parsed_rows.append(
                {
                    "name": name,
                    "email": email,
                    "department": dep,
                    "password": password,
                    "branch": branch,
                }
            )

    # Keep deterministic order and unique email rows (first occurrence wins).
    seen = set()
    unique_rows = []
    for row in parsed_rows:
        email = row.get("email")
        if email in seen:
            continue
        seen.add(email)
        unique_rows.append(row)
    return unique_rows


@app.route("/api/users/bulk-counselors", methods=["POST"])
@login_required
@admin_required
def api_bulk_create_counselors():
    system_only = _ensure_system_admin("users")
    if system_only:
        return system_only

    upload = request.files.get("counsellor_file")
    if not upload or not upload.filename:
        flash("Select a counsellor Excel file to upload.", "error")
        return _redirect_admin_back("users")

    year_level = request.form.get("year_level", type=int) or 1
    if year_level not in (1, 2, 3, 4):
        flash("Year must be between 1 and 4.", "error")
        return _redirect_admin_back("users")

    override_password = str(request.form.get("override_password") or "").strip()
    if override_password and len(override_password) < 6:
        flash("Override password must be at least 6 characters.", "error")
        return _redirect_admin_back("users")

    departments = db.get_departments(active_only=False)
    department_codes = [str(d.get("code") or "").strip().upper() for d in departments if d.get("code")]

    try:
        rows = _parse_bulk_counselor_excel(upload, department_codes)
    except Exception as e:
        flash(f"Could not parse counsellor file: {e}", "error")
        return _redirect_admin_back("users")

    if not rows:
        flash("No valid counselor rows found in the uploaded file.", "warning")
        return _redirect_admin_back("users")

    created = 0
    updated = 0
    skipped = 0
    skipped_emails = []

    for row in rows:
        email = row["email"]
        role_password = override_password or row.get("password") or ""
        if not role_password:
            skipped += 1
            skipped_emails.append(email)
            continue

        if len(role_password) < 6:
            skipped += 1
            skipped_emails.append(email)
            continue

        existing = db.get_user(email)
        if not existing:
            ok, _ = db.create_user(
                email=email,
                password=role_password,
                name=row["name"],
                role="counselor",
                department=row["department"],
                max_students=30,
                can_upload_students=True,
                year_level=year_level,
            )
            if ok:
                created += 1
            else:
                skipped += 1
                skipped_emails.append(email)
            continue

        existing_role = str(existing.get("role") or "").strip().lower()
        if existing_role not in {"counselor"}:
            skipped += 1
            skipped_emails.append(email)
            continue

        db.update_user(
            email,
            name=row["name"],
            role="counselor",
            department=row["department"],
            year_level=year_level,
            can_upload_students=1,
            max_students=30,
            password=role_password,
        )
        updated += 1

    msg = f"Bulk counselor sync completed. Created: {created}, Updated: {updated}, Skipped: {skipped}."
    if skipped_emails:
        msg += f" Skipped emails: {', '.join(skipped_emails[:8])}"
    flash(msg, "success" if (created or updated) else "warning")
    return _redirect_admin_back("users")

@app.route("/api/users", methods=["POST"])
@login_required
@admin_required
def api_create_user():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    email = request.form.get("email", "").strip()
    password = request.form.get("account_password") or request.form.get("password", "")
    confirm_password = request.form.get("account_confirm_password") or request.form.get("confirm_password", "")
    name = request.form.get("name", "").strip()
    role = (request.form.get("role", "counselor") or "counselor").strip().lower()
    if role == "chief_admin":
        role = "hod"
    role = role if role in {"admin", "hod", "deo", "principal", "counselor"} else "counselor"
    year_level = request.form.get("year_level", type=int) or 1

    if _is_hod(actor_role) and role not in {"counselor", "deo"}:
        flash("HoD can create only counselor or DEO accounts.", "error")
        return _redirect_admin_back("users")

    if _is_deo(actor_role) and role != "counselor":
        flash("DEO can create only counselor accounts.", "error")
        return _redirect_admin_back("users")

    if role == "admin" and not _is_system_admin(actor_role):
        flash("Only system admin can create system admin accounts.", "error")
        return _redirect_admin_back("users")

    if role == "principal" and not _is_system_admin(actor_role):
        flash("Only system admin can create principal accounts.", "error")
        return _redirect_admin_back("users")

    scope_pairs = []

    if role == "admin":
        # Admins get unrestricted defaults and do not need counselor-specific fields.
        department = ""
        year_level = 1
        max_students = 500
        can_upload = True
    elif role in {"hod", "deo"}:
        scope_pairs = _parse_scope_pairs_from_form("scope_pairs")

        # Backward-compatible fallback for older forms.
        if not scope_pairs:
            fallback_dep = (request.form.get("department") or "").strip().upper()
            fallback_year = request.form.get("year_level", type=int)
            if fallback_dep and fallback_year in (1, 2, 3, 4):
                scope_pairs = [(fallback_dep, fallback_year)]

        if not scope_pairs:
            flash("Assign at least one department/year scope for HoD/DEO accounts.", "error")
            return _redirect_admin_back("users")

        if _is_hod(actor_role):
            actor_scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
            if not all((dep, yr) in actor_scopes for dep, yr in scope_pairs):
                flash("You can assign only department/year scopes within your own assignment.", "error")
                return _redirect_admin_back("users")

        department, year_level = scope_pairs[0]
        max_students = 500
        can_upload = True
    elif role == "principal":
        department = ""
        year_level = 1
        max_students = 500
        can_upload = True
    else:
        department = request.form.get("department", "").strip()
        max_students_raw = request.form.get("max_students", "30")
        can_upload = True

        try:
            max_students = int(max_students_raw)
        except (TypeError, ValueError):
            flash("Max students must be a valid number.", "error")
            return _redirect_admin_back("users")

        if max_students < 1 or max_students > 500:
            flash("Max students must be between 1 and 500.", "error")
            return _redirect_admin_back("users")

    if not email or not password or not name:
        flash("All required fields must be filled.", "error")
        return _redirect_admin_back("users")

    if password != confirm_password:
        flash("Password and confirm password do not match.", "error")
        return _redirect_admin_back("users")

    if len(password) < 6:
        flash("Password must be at least 6 characters.", "error")
        return _redirect_admin_back("users")

    if role == "counselor" and year_level not in (1, 2, 3, 4):
        flash("Year must be between 1 and 4.", "error")
        return _redirect_admin_back("users")

    if _is_hod(actor_role) and role in {"counselor", "deo"}:
        scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
        if (str(department or "").upper(), int(year_level or 1)) not in scopes:
            flash("You can only create users inside your assigned department/year scope.", "error")
            return _redirect_admin_back("users")

    if _is_deo(actor_role) and role == "counselor":
        scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
        if (str(department or "").upper(), int(year_level or 1)) not in scopes:
            flash("You can only create counselors inside your assigned department/year scope.", "error")
            return _redirect_admin_back("users")

    ok, msg = db.create_user(
        email, password, name, role, department, max_students, can_upload, year_level
    )

    if ok and role == "admin":
        flags = _get_security_config_flags()
        should_disable_default = bool(flags.get("disable_default_admin_on_new_system_admin"))
        default_admin_email = _get_default_admin_email()
        created_email = str(email or "").strip().lower()
        if should_disable_default and default_admin_email and created_email != default_admin_email:
            default_admin = db.get_user(default_admin_email)
            if default_admin:
                db.update_user(
                    default_admin_email,
                    is_active=0,
                    is_locked=1,
                    lock_reason="Disabled automatically after another system admin was created",
                )
                db.force_logout_user(default_admin_email, "default_admin_auto_disabled")
                flash("Default system admin was disabled by policy.", "warning")

    if ok and role in {"hod", "deo"}:
        db.set_chief_admin_scopes(email, scope_pairs or [(department, year_level)])

    # Optional student file during registration (counselors only)
    if ok and role == "counselor" and "student_file" in request.files:
        f = request.files["student_file"]
        if f and f.filename:
            try:
                from core.dynamic_parser import parse_student_excel
                parsed = parse_student_excel(f)
                if parsed:
                    added = db.add_students_bulk(email, parsed)
                    flash(f"User created — {added} students uploaded.", "success")
                else:
                    flash("User created but no valid students found in the uploaded file.", "warning")
            except Exception as e:
                flash(f"User created but student upload failed: {e}", "warning")
            return _redirect_admin_back("users")

    flash(msg, "success" if ok else "error")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/update", methods=["POST"])
@login_required
@admin_required
def api_update_user(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    target = db.get_user(email)
    if not target:
        flash("User not found.", "error")
        return _redirect_admin_back("users")

    if _is_hod(actor_role):
        if target.get("role") not in {"counselor", "deo"}:
            flash("You can modify only counselor or DEO accounts.", "error")
            return _redirect_admin_back("users")

        actor_scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
        target_key = (
            str(target.get("department") or "").strip().upper(),
            int(target.get("year_level") or 1),
        )
        if target_key not in actor_scopes:
            flash("You can modify only users in your assigned scope.", "error")
            return _redirect_admin_back("users")

        requested_dep = (request.form.get("department") or target.get("department") or "").strip().upper()
        requested_year = request.form.get("year_level", type=int) or int(target.get("year_level") or 1)
        requested_key = (requested_dep, requested_year)
        if requested_key not in actor_scopes:
            flash("Update rejected: target department/year is outside your authorized assignments.", "error")
            return _redirect_admin_back("users")

    if _is_deo(actor_role):
        if not _can_deo_touch_user(actor_email, target):
            flash("You can modify only counselors in your assigned scope.", "error")
            return _redirect_admin_back("users")

    updates = {}
    name = request.form.get("name", "").strip()
    if name:
        updates["name"] = name

    requested_role = str(target.get("role") or "counselor").strip().lower() or "counselor"
    requested_scope_pairs = []
    if _is_system_admin(actor_role):
        role_input = str(request.form.get("role") or requested_role).strip().lower()
        if role_input == "chief_admin":
            role_input = "hod"
        if role_input not in {"admin", "hod", "principal", "counselor", "deo"}:
            flash("Invalid role selected.", "error")
            return _redirect_admin_back("users")
        requested_role = role_input
        updates["role"] = requested_role

    password = (request.form.get("account_password") or request.form.get("password", "")).strip()
    if password:
        if len(password) < 6:
            flash("Password must be at least 6 characters.", "error")
            return _redirect_admin_back("users")
        updates["password"] = password

    if _is_system_admin(actor_role):
        if requested_role == "admin":
            updates["department"] = ""
            updates["year_level"] = 1
            updates["max_students"] = 500
            updates["can_upload_students"] = 1
        elif requested_role in {"hod", "deo"}:
            requested_scope_pairs = _parse_scope_pairs_from_form("scope_pairs")

            # Backward-compatible fallback when old single-scope fields are still posted.
            if not requested_scope_pairs:
                fallback_dep = (request.form.get("department") or "").strip().upper()
                fallback_year = request.form.get("year_level", type=int)
                if fallback_dep and fallback_year in (1, 2, 3, 4):
                    requested_scope_pairs = [(fallback_dep, fallback_year)]

            if not requested_scope_pairs:
                flash("Assign at least one department/year scope for HoD/DEO accounts.", "error")
                return _redirect_admin_back("users")

            department, year_level = requested_scope_pairs[0]
            updates["department"] = department
            updates["year_level"] = year_level
            updates["max_students"] = 500
            updates["can_upload_students"] = 1
        elif requested_role == "principal":
            updates["department"] = ""
            updates["year_level"] = 1
            updates["max_students"] = 500
            updates["can_upload_students"] = 1
        else:
            department = (request.form.get("department") or "").strip().upper()
            if not department:
                flash("Department is required for counselor accounts.", "error")
                return _redirect_admin_back("users")
            year_level = request.form.get("year_level", type=int)
            if year_level not in (1, 2, 3, 4):
                flash("Year must be between 1 and 4.", "error")
                return _redirect_admin_back("users")

            max_students_raw = request.form.get("max_students", "")
            try:
                max_students = int(max_students_raw)
            except (TypeError, ValueError):
                flash("Max students must be a valid number.", "error")
                return _redirect_admin_back("users")

            if max_students < 1 or max_students > 500:
                flash("Max students must be between 1 and 500.", "error")
                return _redirect_admin_back("users")

            updates["department"] = department
            updates["year_level"] = year_level
            updates["max_students"] = max_students
            updates["can_upload_students"] = 1 if request.form.get("can_upload_students") == "on" else 0
    else:
        department = (request.form.get("department") or target.get("department") or "").strip().upper()
        year_level = request.form.get("year_level", type=int) or int(target.get("year_level") or 1)
        requested_key = (department, year_level)

        actor_scopes = _get_actor_scope_pairs(actor_email, actor_role) or set()
        if requested_key not in actor_scopes:
            flash("Update rejected: target department/year is outside your authorized assignments.", "error")
            return _redirect_admin_back("users")

        max_students_raw = request.form.get("max_students")
        if max_students_raw:
            try:
                max_students = int(max_students_raw)
            except (TypeError, ValueError):
                flash("Max students must be a valid number.", "error")
                return _redirect_admin_back("users")
            if max_students < 1 or max_students > 500:
                flash("Max students must be between 1 and 500.", "error")
                return _redirect_admin_back("users")
            updates["max_students"] = max_students

        updates["can_upload_students"] = 1 if request.form.get("can_upload_students") == "on" else 0

    db.update_user(email, **updates)

    if _is_system_admin(actor_role):
        if requested_role in {"hod", "deo"}:
            if requested_scope_pairs:
                db.set_chief_admin_scopes(email, requested_scope_pairs)
            else:
                dep = (updates.get("department") if "department" in updates else target.get("department") or "").strip().upper()
                yr = int(updates.get("year_level") if "year_level" in updates else target.get("year_level") or 1)
                db.set_chief_admin_scopes(email, [(dep, yr)])
        else:
            db.set_chief_admin_scopes(email, [])

    flash("User updated.", "success")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/delete", methods=["POST"])
@login_required
@admin_required
def api_delete_user(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    target = db.get_user(email)
    if not target:
        flash("User not found.", "error")
        return _redirect_admin_back("users")
    if _is_hod(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can delete only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    if _is_deo(actor_role) and not _can_deo_touch_user(actor_email, target):
        flash("You can delete only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    try:
        db.delete_user(email)
    except sqlite3.IntegrityError as exc:
        flash(f"User delete blocked by related records: {exc}", "error")
        return _redirect_admin_back("users")
    except Exception as exc:
        flash(f"Unable to delete user right now: {exc}", "error")
        return _redirect_admin_back("users")

    flash("User deleted.", "success")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/lock", methods=["POST"])
@login_required
@admin_required
def api_lock_user(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    target = db.get_user(email)
    if _is_hod(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can lock only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    if _is_deo(actor_role) and not _can_deo_touch_user(actor_email, target):
        flash("You can lock only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    db.lock_user(email, request.form.get("reason", "Locked by admin"))
    flash("User locked.", "success")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/unlock", methods=["POST"])
@login_required
@admin_required
def api_unlock_user(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    target = db.get_user(email)
    if _is_hod(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can unlock only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    if _is_deo(actor_role) and not _can_deo_touch_user(actor_email, target):
        flash("You can unlock only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    db.unlock_user(email)
    flash("User unlocked.", "success")
    return _redirect_admin_back("users")


@app.route("/api/password-update", methods=["POST"])
@login_required
def api_update_password():
    """Update password for logged in user."""
    user_email = session.get("user_email")
    role = session.get("role")
    if not user_email:
        flash("Session expired. Please login again.", "error")
        return redirect(url_for("login"))
    
    current_password = request.form.get("current_password", "").strip()
    new_password = request.form.get("new_password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()
    otp_code = request.form.get("otp_code", "").strip()
    
    # Validate inputs
    if not new_password or not confirm_password:
        flash("New password and confirm password are required.", "error")
        return redirect(request.referrer or url_for(_panel_endpoint_for_role(session.get("role"))))
    
    if len(new_password) < 6:
        flash("New password must be at least 6 characters.", "error")
        return redirect(request.referrer or url_for(_panel_endpoint_for_role(session.get("role"))))
    
    if new_password != confirm_password:
        flash("Passwords do not match.", "error")
        return redirect(request.referrer or url_for(_panel_endpoint_for_role(session.get("role"))))

    user = db.get_user(user_email) or {}

    if _is_password_reset_otp_required_for_user(role, user_email):
        pending = session.get("self_reset_otp") or {}
        expected_email = str(pending.get("email") or "").strip().lower()
        expires_at = _parse_iso_datetime(pending.get("expires_at"))
        if expected_email != str(user_email).strip().lower():
            flash("Request OTP first before updating password.", "error")
            return redirect(request.referrer or url_for(_panel_endpoint_for_role(role)))
        if not expires_at or datetime.now() > expires_at:
            session.pop("self_reset_otp", None)
            flash("OTP expired. Request a new OTP.", "error")
            return redirect(request.referrer or url_for(_panel_endpoint_for_role(role)))
        if not otp_code or _otp_hash(otp_code) != str(pending.get("otp_hash") or ""):
            flash("Invalid OTP.", "error")
            return redirect(request.referrer or url_for(_panel_endpoint_for_role(role)))
        session.pop("self_reset_otp", None)
    else:
        # Keep current-password verification mandatory for system admin.
        if _is_system_admin(role):
            if not current_password:
                flash("Current password is required.", "error")
                return redirect(request.referrer or url_for(_panel_endpoint_for_role(role)))
            stored_hash = user.get("password_hash")
            if not db.verify_password(current_password, stored_hash):
                flash("Current password is incorrect.", "error")
                return redirect(request.referrer or url_for(_panel_endpoint_for_role(role)))

    # Update password
    db.update_user(user_email, password=new_password)
    
    flash("Password updated successfully.", "success")
    return redirect(request.referrer or url_for(_panel_endpoint_for_role(session.get("role"))))


@app.route("/api/password-reset/send-otp", methods=["POST"])
@login_required
def api_send_password_reset_otp():
    user_email = session.get("user_email")
    role = session.get("role")
    if not user_email:
        flash("Session expired. Please login again.", "error")
        return redirect(url_for("login"))

    if not _is_password_reset_otp_required_for_user(role, user_email):
        flash("OTP is not required for your role at this time.", "info")
        return redirect(request.referrer or url_for(_panel_endpoint_for_role(role)))

    pending = session.get("self_reset_otp") or {}
    last_requested = _parse_iso_datetime(pending.get("requested_at"))
    if last_requested and (datetime.now() - last_requested).total_seconds() < 30:
        flash("Please wait 30 seconds before requesting another OTP.", "warning")
        return redirect(request.referrer or url_for(_panel_endpoint_for_role(role)))

    otp_code = generate_otp(6)
    if not _send_otp_email(user_email, otp_code, "Password Reset"):
        flash("Unable to send OTP. Check mail configuration.", "error")
        return redirect(request.referrer or url_for(_panel_endpoint_for_role(role)))

    session["self_reset_otp"] = {
        "email": user_email,
        "otp_hash": _otp_hash(otp_code),
        "expires_at": (datetime.now() + timedelta(seconds=OTP_EXPIRY_SECONDS)).isoformat(),
        "requested_at": datetime.now().isoformat(),
    }
    session.modified = True
    flash(f"OTP sent to {_mask_email(user_email)}.", "success")
    return redirect(request.referrer or url_for(_panel_endpoint_for_role(role)))


@app.route("/api/users/<path:email>/upload-students", methods=["POST"])
@login_required
@admin_required
def api_upload_students_for_counselor(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    target = db.get_user(email)
    if _is_hod(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can upload students only for counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    if _is_deo(actor_role) and not _can_deo_touch_user(actor_email, target):
        flash("You can upload students only for counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    f = request.files.get("student_file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return _redirect_admin_back("users")
    try:
        from core.dynamic_parser import parse_student_excel
        parsed = parse_student_excel(f)
        if parsed:
            added = db.add_students_bulk(email, parsed)
            flash(f"{added} students uploaded for {email}.", "success")
        else:
            flash("No valid students found.", "error")
    except Exception as e:
        flash(f"Upload failed: {e}", "error")
    return _redirect_admin_back("users")


@app.route("/api/users/<path:email>/force-logout", methods=["POST"])
@login_required
@admin_required
def api_force_logout(email):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    target = db.get_user(email)
    if _is_hod(actor_role) and not _can_chief_admin_touch_user(actor_email, target):
        flash("You can force logout only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    if _is_deo(actor_role) and not _can_deo_touch_user(actor_email, target):
        flash("You can force logout only counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    db.force_logout_user(email, "admin_action")
    flash(f"Force-logged-out {email}.", "success")
    return _redirect_admin_back("users")


@app.route("/api/users/reset-password", methods=["POST"])
@login_required
@admin_required
def api_admin_reset_password():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_deo(actor_role) or _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    target_email = request.form.get("target_email", "").strip()
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")
    force_logout = request.form.get("force_logout") == "on"

    if not target_email or not new_password or not confirm_password:
        flash("User and both password fields are required.", "error")
        return redirect(url_for(_panel_endpoint_for_role(actor_role), tab="config"))

    if new_password != confirm_password:
        flash("New password and confirm password do not match.", "error")
        return redirect(url_for(_panel_endpoint_for_role(actor_role), tab="config"))

    if len(new_password) < 6:
        flash("Password must be at least 6 characters.", "error")
        return redirect(url_for(_panel_endpoint_for_role(actor_role), tab="config"))

    user = db.get_user(target_email)
    if not user:
        flash("Selected user was not found.", "error")
        return redirect(url_for(_panel_endpoint_for_role(actor_role), tab="config"))

    if _is_hod(actor_role) and not _can_chief_admin_touch_user(actor_email, user):
        flash("You can reset passwords only for counselors in your assigned scope.", "error")
        return redirect(url_for(_panel_endpoint_for_role(actor_role), tab="users"))

    db.update_user(target_email, password=new_password)

    # Security best-practice: invalidate existing sessions after password reset.
    if force_logout and target_email != session.get("user_email"):
        db.force_logout_user(target_email, "admin_password_reset")

    flash(f"Password updated successfully for {target_email}.", "success")
    return redirect(url_for(_panel_endpoint_for_role(actor_role), tab="config"))


@app.route("/api/chief-admin/reset-counselor-password", methods=["POST"])
@login_required
@admin_required
def api_chief_admin_reset_password():
    """HoD can reset password for counselors/DEOs in their assigned dept/year scope."""
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    
    if not _is_hod(actor_role):
        flash("Only HoD users can access this function.", "error")
        return redirect(url_for("counselor_page"))
    
    target_email = request.form.get("target_email", "").strip()
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")
    force_logout = request.form.get("force_logout") == "on"

    if not target_email or not new_password or not confirm_password:
        flash("Target user and both password fields are required.", "error")
        return redirect(url_for("counselor_page"))

    if new_password != confirm_password:
        flash("New password and confirm password do not match.", "error")
        return redirect(url_for("counselor_page"))

    if len(new_password) < 6:
        flash("Password must be at least 6 characters.", "error")
        return redirect(url_for("counselor_page"))

    target = db.get_user(target_email)
    if not target:
        flash("Selected user was not found.", "error")
        return redirect(url_for("counselor_page"))
    
    if target.get("role") not in {"counselor", "deo"}:
        flash("You can reset passwords only for counselors or DEOs.", "error")
        return redirect(url_for("counselor_page"))

    # HoD must be able to manage this target's department/year
    if not _can_chief_admin_touch_user(actor_email, target):
        flash("You can reset passwords only for users in your assigned scope.", "error")
        return redirect(url_for("counselor_page"))

    db.update_user(target_email, password=new_password)

    # Security best-practice: invalidate existing sessions after password reset.
    if force_logout and target_email != session.get("user_email"):
        db.force_logout_user(target_email, "hod_password_reset")

    flash(f"Password updated successfully for {target_email}.", "success")
    return redirect(url_for("counselor_page"))


@app.route("/api/chief-admin/scoped-counselors", methods=["GET"])
@login_required
@admin_required
def api_chief_admin_get_scoped_counselors():
    """Return list of counselors/DEOs under HoD's dept/year scope as JSON."""
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    
    if not _is_hod(actor_role):
        return jsonify({"error": "Unauthorized"}), 403
    
    # Get all users scoped to this HoD
    all_scoped_users = db.get_scoped_users_for_admin(actor_email, actor_role) or []
    
    # Filter to counselors + DEOs
    counselors = [
        {
            "name": u.get("name", ""),
            "email": u.get("email", ""),
            "department": u.get("department", ""),
            "year_level": u.get("year_level", 1)
        }
        for u in all_scoped_users
        if u.get("role") in {"counselor", "deo"}
    ]
    
    return jsonify(counselors)


@app.route("/api/admin/students/save", methods=["POST"])
@login_required
@admin_required
def api_admin_save_student():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    counselor_email = request.form.get("counselor_email", "").strip()
    original_reg_no = request.form.get("original_reg_no", "").strip()
    reg_no = request.form.get("reg_no", "").strip()
    student_name = request.form.get("student_name", "").strip()
    parent_phone = request.form.get("parent_phone", "").strip()

    if not counselor_email or not reg_no or not student_name:
        flash("Counselor, Register No and Student Name are required.", "error")
        return _redirect_admin_back("users", open_manage=counselor_email)

    counselor = db.get_user(counselor_email)
    if not counselor or counselor.get("role") != "counselor":
        flash("Invalid counselor selected.", "error")
        return _redirect_admin_back("users")

    if _is_hod(actor_role) and not _can_chief_admin_touch_user(actor_email, counselor):
        flash("You can manage students only for counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    if _is_deo(actor_role) and not _can_deo_touch_user(actor_email, counselor):
        flash("You can manage students only for counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")

    department = (counselor.get("department") or "").strip().upper()

    try:
        if original_reg_no and original_reg_no != reg_no:
            db.delete_student(counselor_email, original_reg_no)

        db.admin_upsert_student(
            counselor_email,
            reg_no,
            student_name,
            department=department,
            parent_phone=parent_phone,
        )
        flash(f"Student {reg_no} saved for {counselor.get('name')}", "success")
    except Exception as e:
        flash(f"Could not save student: {e}", "error")

    return _redirect_admin_back("users", open_manage=counselor_email)


@app.route("/api/admin/students/delete", methods=["POST"])
@login_required
@admin_required
def api_admin_delete_student():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    counselor_email = request.form.get("counselor_email", "").strip()
    reg_no = request.form.get("reg_no", "").strip()

    if not counselor_email or not reg_no:
        flash("Counselor and Register No are required.", "error")
        return _redirect_admin_back("users", open_manage=counselor_email)

    counselor = db.get_user(counselor_email)
    if not counselor or counselor.get("role") != "counselor":
        flash("Invalid counselor selected.", "error")
        return _redirect_admin_back("users")

    if _is_hod(actor_role) and not _can_chief_admin_touch_user(actor_email, counselor):
        flash("You can manage students only for counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    if _is_deo(actor_role) and not _can_deo_touch_user(actor_email, counselor):
        flash("You can manage students only for counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")

    try:
        db.delete_student(counselor_email, reg_no)
        flash(f"Deleted student {reg_no}.", "success")
    except Exception as e:
        flash(f"Could not delete student: {e}", "error")

    return _redirect_admin_back("users", open_manage=counselor_email)


@app.route("/api/admin/students/delete-all", methods=["POST"])
@login_required
@admin_required
def api_admin_delete_all_students():
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        flash("You do not have permission to manage users.", "error")
        return _redirect_admin_back("reports")
    counselor_email = request.form.get("counselor_email", "").strip()
    if not counselor_email:
        flash("Counselor is required.", "error")
        return _redirect_admin_back("users")

    counselor = db.get_user(counselor_email)
    if not counselor or counselor.get("role") != "counselor":
        flash("Invalid counselor selected.", "error")
        return _redirect_admin_back("users")

    if _is_hod(actor_role) and not _can_chief_admin_touch_user(actor_email, counselor):
        flash("You can manage students only for counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")
    if _is_deo(actor_role) and not _can_deo_touch_user(actor_email, counselor):
        flash("You can manage students only for counselors in your assigned scope.", "error")
        return _redirect_admin_back("users")

    try:
        db.delete_all_students(counselor_email)
        flash(f"Deleted all students for {counselor.get('name')}", "success")
    except Exception as e:
        flash(f"Could not delete student list: {e}", "error")

    return _redirect_admin_back("users", open_manage=counselor_email)


# ---------- Departments -----------------------------------------------------

@app.route("/api/departments", methods=["POST"])
@login_required
@admin_required
def api_create_department():
    if not _is_system_admin(session.get("role")):
        flash("Only system admin can create departments.", "error")
        return _redirect_admin_back("departments")
    code = request.form.get("code", "").strip().upper()
    name = request.form.get("name", "").strip()
    color = request.form.get("color", "#667eea")
    if not code or not name:
        flash("Code and name are required.", "error")
        return _redirect_admin_back("departments")
    ok, msg = db.create_department(code, name, color)
    flash(msg, "success" if ok else "error")
    return _redirect_admin_back("departments")


@app.route("/api/departments/<int:dept_id>/delete", methods=["POST"])
@login_required
@admin_required
def api_delete_department(dept_id):
    if not _is_system_admin(session.get("role")):
        flash("Only system admin can delete departments.", "error")
        return _redirect_admin_back("departments")
    db.delete_department(dept_id)
    flash("Department deleted.", "success")
    return _redirect_admin_back("departments")


@app.route("/api/departments/<int:dept_id>/update", methods=["POST"])
@login_required
@admin_required
def api_update_department(dept_id):
    if not _is_system_admin(session.get("role")):
        flash("Only system admin can edit departments.", "error")
        return _redirect_admin_back("departments")

    code = request.form.get("code", "").strip().upper()
    name = request.form.get("name", "").strip()
    if not code or not name:
        flash("Department code and full name are required.", "error")
        return _redirect_admin_back("departments")

    ok, msg = db.update_department_identity(dept_id, code, name)
    flash(msg, "success" if ok else "error")
    return _redirect_admin_back("departments")


@app.route("/api/departments/<int:dept_id>/toggle", methods=["POST"])
@login_required
@admin_required
def api_toggle_department(dept_id):
    actor_role = session.get("role")
    if not _is_system_admin(actor_role):
        flash("Only system admin can enable or disable departments.", "error")
        return _redirect_admin_back("departments")

    is_active = request.form.get("is_active") == "1"
    db.update_department(dept_id, is_active=0 if is_active else 1)
    flash("Department updated.", "success")
    return _redirect_admin_back("departments")


# ---------- Sessions --------------------------------------------------------

@app.route("/api/sessions/cleanup", methods=["POST"])
@login_required
@admin_required
def api_cleanup_sessions():
    system_only = _ensure_system_admin("monitoring")
    if system_only:
        return system_only
    db.cleanup_stale_sessions()
    db.clear_inactive_sessions()
    flash("Sessions cleaned.", "success")
    return _redirect_admin_back("monitoring")


@app.route("/api/sessions/logout-all", methods=["POST"])
@login_required
@admin_required
def api_logout_all():
    system_only = _ensure_system_admin("monitoring")
    if system_only:
        return system_only
    db.logout_all_users()
    flash("All users logged out.", "success")
    return _redirect_admin_back("monitoring")


# ---------- App Configuration -----------------------------------------------

@app.route("/api/config/update", methods=["POST"])
@login_required
@admin_required
def api_update_config():
    system_only = _ensure_system_admin("config")
    if system_only:
        return system_only

    settings = {}
    
    # Session timeout (hours -> seconds)
    timeout = request.form.get("session_timeout")
    if timeout:
        try:
            timeout_hours = int(timeout)
            if timeout_hours < 1 or timeout_hours > 168:
                flash("Session timeout must be between 1 and 168 hours.", "error")
                return _redirect_admin_back("config")
            settings["session_timeout"] = str(timeout_hours * 3600)
        except ValueError:
            flash("Invalid session timeout value.", "error")
            return _redirect_admin_back("config")
    
    # Heartbeat interval
    heartbeat = request.form.get("session_heartbeat_interval")
    if heartbeat:
        try:
            settings["session_heartbeat_interval"] = str(int(heartbeat))
        except ValueError:
            pass
    
    # Hex color settings
    color_fields = [
        "color_primary", "color_primary_dark", "color_secondary", "color_accent",
        "color_success", "color_warning", "color_danger", "color_info",
        "color_bg_primary", "color_bg_secondary",
        "color_text", "color_text_dim", "color_text_muted"
    ]
    for field in color_fields:
        value = request.form.get(field)
        if value and value.startswith("#"):
            settings[field] = value

    # Advanced color fields that may use rgba()/hex formats
    advanced_color_fields = ["color_bg_card", "color_border"]
    for field in advanced_color_fields:
        value = request.form.get(field)
        if value:
            settings[field] = value.strip()
    
    # Session monitoring settings
    session_monitoring = request.form.get("session_monitoring_enabled")
    settings["session_monitoring_enabled"] = "true" if session_monitoring == "on" else "false"
    
    allow_concurrent = request.form.get("allow_concurrent_sessions")
    settings["allow_concurrent_sessions"] = "true" if allow_concurrent == "on" else "false"
    
    max_concurrent = request.form.get("max_concurrent_sessions")
    if max_concurrent:
        settings["max_concurrent_sessions"] = str(max_concurrent)

    tutorial_master = request.form.get("tutorial_master_enabled")
    settings["tutorial_master_enabled"] = "true" if tutorial_master == "on" else "false"

    tutorial_fields = [
        "tutorial_counselor_enabled",
        "tutorial_hod_enabled",
        "tutorial_deo_enabled",
        "tutorial_principal_enabled",
    ]
    for field in tutorial_fields:
        field_value = request.form.get(field)
        settings[field] = "true" if (tutorial_master == "on" and field_value == "on") else "false"

    # System-admin lifecycle policy.
    settings["disable_default_admin_on_new_system_admin"] = "true" if request.form.get("disable_default_admin_on_new_system_admin") == "on" else "false"

    # OTP security toggles
    request_reset_otp = request.form.get("require_otp_on_password_reset") == "on"
    request_login_otp = request.form.get("require_otp_on_login") == "on"
    if request_reset_otp or request_login_otp:
        smtp_state = _resolve_smtp_status(force_refresh=True).get("state")
        smtp_ready = smtp_state in {"ready", "test"}
        if not smtp_ready:
            settings["require_otp_on_password_reset"] = "false"
            settings["require_otp_on_login"] = "false"
            flash("OTP options were turned off because SMTP is not ready.", "warning")
        else:
            settings["require_otp_on_password_reset"] = "true" if request_reset_otp else "false"
            settings["require_otp_on_login"] = "true" if request_login_otp else "false"
    else:
        settings["require_otp_on_password_reset"] = "false"
        settings["require_otp_on_login"] = "false"
    
    if settings:
        db.update_app_config_bulk(settings)
        flash("Configuration updated successfully.", "success")
    
    return _redirect_admin_back("config")


@app.route("/api/config/smtp-refresh", methods=["POST"])
@login_required
@admin_required
def api_config_smtp_refresh():
    system_only = _ensure_system_admin("config")
    if system_only:
        return system_only

    status = _resolve_smtp_status(force_refresh=True)
    if status.get("state") in {"ready", "test"}:
        flash(f"{status.get('label')}: {status.get('detail')}", "success")
    elif status.get("state") == "missing":
        flash(status.get("detail") or "SMTP credentials are missing.", "warning")
    else:
        flash(status.get("detail") or "SMTP check failed.", "error")
    return _redirect_admin_back("config")


@app.route("/api/config/smtp-status", methods=["GET"])
@login_required
@admin_required
def api_config_smtp_status():
    system_only = _ensure_system_admin("config")
    if system_only:
        return jsonify({"success": False, "message": "System admin access required."}), 403

    force_refresh = str(request.args.get("refresh") or "").strip() in {"1", "true", "yes"}
    status = _resolve_smtp_status(force_refresh=force_refresh)
    return jsonify({"success": True, "status": status})


@app.route("/api/config/smtp-test", methods=["POST"])
@login_required
@admin_required
def api_config_smtp_test():
    system_only = _ensure_system_admin("config")
    if system_only:
        return system_only

    status = _resolve_smtp_status(force_refresh=True)
    if status.get("state") not in {"ready", "test"}:
        flash("SMTP is not ready. Refresh SMTP status and fix configuration first.", "error")
        return _redirect_admin_back("config")

    target_email = session.get("user_email") or ""
    if not target_email:
        flash("Unable to resolve your account email for test message.", "error")
        return _redirect_admin_back("config")

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    subject = "RMKCET SHINE SMTP Test"
    body = (
        "<h3>SMTP Test Successful</h3>"
        f"<p>Timestamp: {stamp}</p>"
        f"<p>Server: {SMTP_SERVER}:{SMTP_PORT}</p>"
        "<p>This is a verification mail from RMKCET SHINE.</p>"
    )
    if send_email(target_email, subject, body, html=True):
        flash(f"SMTP test email sent to {_mask_email(target_email)}.", "success")
    else:
        flash("SMTP test email failed. Check credentials and provider security settings.", "error")
    return _redirect_admin_back("config")


@app.route("/api/config/reset-theme", methods=["POST"])
@login_required
@admin_required
def api_reset_theme():
    system_only = _ensure_system_admin("config")
    if system_only:
        return system_only

    # Reset all color settings to defaults
    defaults = {
        "color_primary": "#667eea",
        "color_primary_dark": "#5a6fd6",
        "color_secondary": "#764ba2",
        "color_accent": "#a78bfa",
        "color_success": "#25D366",
        "color_warning": "#f59e0b",
        "color_danger": "#ef4444",
        "color_info": "#3b82f6",
        "color_bg_primary": "#0a0c14",
        "color_bg_secondary": "#0f1219",
        "color_bg_card": "rgba(20, 30, 50, 0.65)",
        "color_text": "#e2e8f0",
        "color_text_dim": "#94a3b8",
        "color_text_muted": "#64748b",
        "color_border": "rgba(102, 126, 234, 0.18)"
    }
    db.update_app_config_bulk(defaults)
    flash("All theme colors reset to defaults.", "success")
    return _redirect_admin_back("config")


# ---------- Activity Export -------------------------------------------------

@app.route("/api/activity/export/csv")
@login_required
@admin_required
def api_export_activity_csv():
    data = _filter_activity_for_actor(
        db.get_counselor_activity_summary(),
        session.get("user_email"),
        session.get("role"),
    )
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Name", "Email", "Department", "Students", "Tests",
                "Messages", "Status", "Last Login"])
    for a in data:
        w.writerow([a["name"], a["email"], a["department"],
                     a["student_count"], a["tests_uploaded"],
                     a["total_messages"], a["work_status"],
                     a["last_login"] or "Never"])
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=counselor_activity.csv"},
    )


@app.route("/api/activity/export/excel")
@login_required
@admin_required
def api_export_activity_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    data = _filter_activity_for_actor(
        db.get_counselor_activity_summary(),
        session.get("user_email"),
        session.get("role"),
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Counselor Activity"

    headers = ["Name", "Email", "Department", "Students", "Tests",
               "Messages", "Week Msgs", "Status", "Last Login"]
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, a in enumerate(data, 2):
        ws.cell(row=ri, column=1, value=a["name"])
        ws.cell(row=ri, column=2, value=a["email"])
        ws.cell(row=ri, column=3, value=a["department"])
        ws.cell(row=ri, column=4, value=a["student_count"])
        ws.cell(row=ri, column=5, value=a["tests_uploaded"])
        ws.cell(row=ri, column=6, value=a["total_messages"])
        ws.cell(row=ri, column=7, value=a["week_messages"])
        ws.cell(row=ri, column=8, value=a["work_status"])
        ws.cell(row=ri, column=9, value=a["last_login"] or "Never")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="counselor_activity.xlsx")


@app.route("/api/activity/export/pdf")
@login_required
@admin_required
def api_export_activity_pdf():
    data = _filter_activity_for_actor(
        db.get_counselor_activity_summary(),
        session.get("user_email"),
        session.get("role"),
    )
    pdf = FPDF("L")
    pdf.add_page()
    pdf.set_font("helvetica", "B", 16)
    pdf.cell(0, 12, "RMKCET Parent Connect - Counselor Activity", 0, 1, "C")
    pdf.set_font("helvetica", "I", 10)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}", 0, 1, "C")
    pdf.ln(4)

    widths = [45, 55, 28, 22, 20, 26, 42, 40]
    heads = ["Name", "Email", "Dept", "Students", "Tests", "Messages", "Status", "Last Login"]
    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(102, 126, 234)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(widths, heads):
        pdf.cell(w, 8, h, 1, 0, "C", True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", "", 8)
    for a in data:
        login = (a["last_login"] or "Never")[:16]
        pdf.cell(widths[0], 7, a["name"][:22], 1)
        pdf.cell(widths[1], 7, a["email"][:28], 1)
        pdf.cell(widths[2], 7, a["department"][:10], 1, 0, "C")
        pdf.cell(widths[3], 7, str(a["student_count"]), 1, 0, "C")
        pdf.cell(widths[4], 7, str(a["tests_uploaded"]), 1, 0, "C")
        pdf.cell(widths[5], 7, str(a["total_messages"]), 1, 0, "C")
        pdf.cell(widths[6], 7, a["work_status"][:22], 1, 0, "C")
        pdf.cell(widths[7], 7, login, 1, 1)

    buf = io.BytesIO()
    raw = pdf.output()
    if isinstance(raw, str):
        raw = raw.encode("latin-1")
    else:
        raw = bytes(raw)
    buf.write(raw)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name="counselor_activity.pdf")


@app.route("/api/activity/export/filtered-pdf")
@login_required
@admin_required
def api_export_activity_filtered_pdf():
    role = session.get("role")
    act_department = (request.args.get("act_dept") or "").strip().upper()
    act_year = request.args.get("act_year", type=int) or 1
    act_sem = (request.args.get("act_sem") or "").strip()
    act_test = _normalize_allowed_test_name(request.args.get("act_test") or "")
    act_q = (request.args.get("act_q") or "").strip()
    act_sort = (request.args.get("act_sort") or "pending_first").strip()

    if _is_hod(role) or _is_deo(role):
        allowed_scopes = {
            (str(s.get("department") or "").strip().upper(), int(s.get("year_level") or 1))
            for s in db.get_chief_admin_scopes(session.get("user_email"))
        }
        req_key = (act_department.strip().upper(), int(act_year or 0))
        if req_key not in allowed_scopes:
            flash("Select an allocated department/year before exporting PDF.", "error")
            return _redirect_admin_back("activity")

    if not (act_department and act_year in (1, 2, 3, 4) and act_sem in {"1", "2"} and act_test):
        flash("Select department, year, semester and test before exporting PDF.", "error")
        return _redirect_admin_back("activity")

    data = db.get_counselor_activity_for_test(
        department=act_department,
        year_level=act_year,
        semester=act_sem,
        test_name=act_test,
        search_query=act_q,
        sort_mode=act_sort,
    )

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    logo_path = os.path.join(FRONTEND_DIR, "static", "assets", "shine-logo.png")
    if os.path.isfile(logo_path):
        pdf.image(logo_path, x=10, y=8, w=18)

    pdf.set_font("helvetica", "B", 16)
    pdf.cell(0, 10, "RMKCET SHINE", align="C")
    pdf.ln()
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(
        0,
        8,
        f"Counselor Activity - {act_department} Year {act_year} Semester {act_sem} - {act_test}",
        align="C",
    )
    pdf.ln()
    pdf.set_font("helvetica", "", 9)
    stats = data.get("stats") or {}
    pdf.cell(
        0,
        6,
        (
            f"Total Counselors: {stats.get('total_counselors', 0)}   "
            f"Complete: {stats.get('complete', 0)}   "
            f"Pending: {stats.get('pending', 0)}   "
            f"Average Completion: {stats.get('avg_completion', 0)}%"
        ),
        align="L",
    )
    pdf.ln()
    pdf.ln(2)

    headers = ["Counselor", "Email", "Students", "Reached", "Pending", "Completion %", "Status", "Last Login"]
    widths = [45, 70, 20, 20, 20, 25, 25, 45]
    pdf.set_font("helvetica", "B", 9)
    for w, h in zip(widths, headers):
        pdf.cell(w, 8, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("helvetica", "", 8)
    for r in data.get("rows") or []:
        pdf.cell(widths[0], 7, str(r.get("name") or "")[:26], border=1)
        pdf.cell(widths[1], 7, str(r.get("email") or "")[:40], border=1)
        pdf.cell(widths[2], 7, str(r.get("student_count") or 0), border=1, align="C")
        pdf.cell(widths[3], 7, str(r.get("unique_students_messaged") or 0), border=1, align="C")
        pdf.cell(widths[4], 7, str(r.get("pending_count") or 0), border=1, align="C")
        pdf.cell(widths[5], 7, f"{int(r.get('completion_pct') or 0)}%", border=1, align="C")
        pdf.cell(widths[6], 7, str(r.get("work_status") or ""), border=1, align="C")
        pdf.cell(widths[7], 7, str((r.get("last_login") or "Never"))[:16], border=1, align="C")
        pdf.ln()

    buf = io.BytesIO()
    raw = pdf.output()
    if isinstance(raw, str):
        raw = raw.encode("latin-1")
    else:
        raw = bytes(raw)
    buf.write(raw)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"activity_{act_department}_Y{act_year}_S{act_sem}_{act_test.replace(' ', '_')}.pdf",
    )


@app.route("/api/activity/<path:email>")
@login_required
@admin_required
def api_activity_detail(email):
    if _is_hod(session.get("role")):
        target = db.get_user(email)
        if not _can_chief_admin_touch_user(session.get("user_email"), target):
            return jsonify({"error": "Access denied"}), 403

    detail = db.get_counselor_detailed_activity(email)
    if not detail:
        return jsonify({"error": "Not found"}), 404
    # Make JSON-serializable
    for key in list(detail.keys()):
        val = detail[key]
        if isinstance(val, datetime):
            detail[key] = val.isoformat()
    return jsonify(detail)


# ---------- Counselor student / marks / reports -----------------------------

@app.route("/api/students/upload", methods=["POST"])
@login_required
def api_upload_students():
    email = session["user_email"]
    user = db.get_user(email)
    if not user.get("can_upload_students", 1):
        flash("You do not have permission to upload students.", "error")
        return redirect(url_for("counselor_page"))

    f = request.files.get("student_file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return redirect(url_for("counselor_page"))

    try:
        from core.dynamic_parser import parse_student_excel
        parsed = parse_student_excel(f)
        if parsed:
            added = db.add_students_bulk(email, parsed)
            flash(f"{added} students uploaded successfully.", "success")
        else:
            flash("No valid students found in the file.", "error")
    except Exception as e:
        flash(f"Upload failed: {e}", "error")
    return redirect(url_for("counselor_page"))


@app.route("/api/students/<reg_no>/delete", methods=["POST"])
@login_required
def api_delete_student(reg_no):
    db.delete_student(session["user_email"], reg_no)
    flash("Student removed.", "success")
    return redirect(url_for("counselor_page"))


@app.route("/api/students/delete-all", methods=["POST"])
@login_required
def api_delete_all_students():
    db.delete_all_students(session["user_email"])
    flash("All students removed.", "success")
    return redirect(url_for("counselor_page"))


# ---------- Tests (Admin) ---------------------------------------------------

@app.route("/api/tests/<int:test_id>/delete", methods=["POST"])
@login_required
@admin_required
def api_delete_test(test_id):
    try:
        actor_email = session.get("user_email")
        actor_role = session.get("role")
        meta = db.get_test_metadata(test_id) or {}
        if not meta:
            flash("Test not found.", "error")
            return _redirect_admin_back("reports")
        if not _can_manage_department_year(actor_email, actor_role, meta.get("department"), meta.get("year_level") or 1):
            flash("You can manage tests only in your assigned department/year scope.", "error")
            return _redirect_admin_back("reports")
        db.delete_test(test_id)
        flash("Test deleted successfully.", "success")
    except Exception as e:
        flash(f"Failed to delete test: {e}", "error")
    return _redirect_admin_back("reports")


@app.route("/api/tests/<int:test_id>/update", methods=["POST"])
@login_required
@admin_required
def api_update_test(test_id):
    try:
        actor_email = session.get("user_email")
        actor_role = session.get("role")
        meta = db.get_test_metadata(test_id) or {}
        if not _can_manage_department_year(actor_email, actor_role, meta.get("department"), meta.get("year_level") or 1):
            flash("You can manage tests only in your assigned department/year scope.", "error")
            return _redirect_admin_back("reports")
        test_name_raw = (request.form.get("test_name") or "").strip() or (meta.get("test_name") or "")
        test_name = _normalize_allowed_test_name(test_name_raw)
        if not test_name:
            flash("Test name must be one of: IAT 1, IAT 2, MODEL EXAM.", "error")
            return _redirect_admin_back("reports")
        semester = (request.form.get("semester") or "").strip() or (meta.get("semester") or "")
        department = (request.form.get("department") or "").strip() or (meta.get("department") or "")
        batch_name = (request.form.get("batch_name") or "").strip() or (meta.get("batch_name") or "")
        section = (request.form.get("section") or "").strip() or (meta.get("section") or "")

        db.update_test_metadata_fields(
            test_id,
            test_name=test_name,
            semester=semester,
            department=department,
            batch_name=batch_name,
            section=section,
        )
        flash("Test updated successfully.", "success")
    except Exception as e:
        flash(f"Failed to update test: {e}", "error")
    return _redirect_admin_back("reports")


@app.route("/api/tests/<int:test_id>/toggle-block", methods=["POST"])
@login_required
@admin_required
def api_toggle_test_block(test_id):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    meta = db.get_test_metadata(test_id) or {}
    if not meta:
        flash("Test not found.", "error")
        return _redirect_admin_back("reports")

    can_toggle = _is_principal(actor_role) or _can_manage_department_year(
        actor_email,
        actor_role,
        meta.get("department"),
        meta.get("year_level") or 1,
    )
    if not can_toggle:
        flash("You can manage tests only in your assigned department/year scope.", "error")
        return _redirect_admin_back("reports")

    current = int(meta.get("is_blocked") or 0)
    next_value = 0 if current else 1
    db.update_test_block_status(test_id, next_value)
    flash("Test blocked." if next_value else "Test unblocked.", "success")
    return _redirect_admin_back(
        "reports",
        report_dept=(meta.get("department") or "").strip().upper(),
        report_year=int(meta.get("year_level") or 1),
    )


@app.route("/admin/tests/<int:test_id>/edit")
@login_required
@admin_required
def admin_test_edit_page(test_id):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    meta = db.get_test_metadata(test_id) or {}
    if not meta:
        flash("Test not found.", "error")
        return _redirect_admin_back("reports")
    if not _can_manage_department_year(actor_email, actor_role, meta.get("department"), meta.get("year_level") or 1):
        flash("Access denied for this test.", "error")
        return _redirect_admin_back("reports")

    grouped = db.get_test_marks_grouped(test_id)
    return render_template(
        "admin_test_edit.html",
        test_id=test_id,
        test_meta=meta,
        subjects=grouped.get("subjects", []),
        students=grouped.get("students", []),
        is_read_only=_is_principal(actor_role),
    )


@app.route("/api/tests/<int:test_id>/marks/update", methods=["POST"])
@login_required
@admin_required
def api_admin_update_marks(test_id):
    actor_email = session.get("user_email")
    actor_role = session.get("role")
    if _is_principal(actor_role):
        return jsonify({"success": False, "error": "Principal is read-only."}), 403

    meta = db.get_test_metadata(test_id) or {}
    if not meta:
        return jsonify({"success": False, "error": "Test not found."}), 404
    if not _can_manage_department_year(actor_email, actor_role, meta.get("department"), meta.get("year_level") or 1):
        return jsonify({"success": False, "error": "Access denied."}), 403

    payload = request.get_json(force=True, silent=False) or {}
    reg_no = str(payload.get("reg_no") or "").strip()
    marks = payload.get("marks") or {}
    if not reg_no or not isinstance(marks, dict):
        return jsonify({"success": False, "error": "Invalid payload."}), 400

    try:
        for subject_name, mark_value in marks.items():
            if not str(subject_name or "").strip():
                continue
            db.upsert_test_mark(
                test_id=test_id,
                reg_no=reg_no,
                subject_name=str(subject_name),
                marks=str(mark_value or ""),
                department=meta.get("department") or "",
                uploaded_by=actor_email,
            )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/admin/tests/upload", methods=["POST"])
@login_required
@admin_required
def api_admin_upload_marksheet():
    role = session.get("role")
    if _is_hod(role) or _is_principal(role):
        flash("You do not have permission to upload marksheets.", "error")
        return _redirect_admin_back("reports")

    f = request.files.get("marks_file")
    if not f or not f.filename:
        flash("No marks file selected.", "error")
        return _redirect_admin_back("reports")

    department = (request.form.get("department") or "").strip().upper()
    year_level = request.form.get("year_level", type=int) or 1
    semester = (request.form.get("semester") or "").strip()
    batch_name = (request.form.get("batch_name") or "").strip()
    section = (request.form.get("section") or "").strip()
    test_name_input = (request.form.get("test_name") or "").strip()
    upload_mode = (request.form.get("upload_mode") or "new").strip().lower()

    if not department or year_level not in (1, 2, 3, 4) or not semester:
        flash("Department, year and semester are required.", "error")
        return _redirect_admin_back("reports", report_dept=department, report_year=year_level)

    if not _can_manage_department_year(session.get("user_email"), session.get("role"), department, year_level):
        flash("You can upload tests only in your assigned department/year scope.", "error")
        return _redirect_admin_back("reports", report_dept=department, report_year=year_level)

    try:
        file_bytes = f.read()
        file_hash = hashlib.sha256(file_bytes).hexdigest()

        from core.intelligent_parser import IntelligentParser
        parser = IntelligentParser()
        test_info, students = parser.parse_file(io.BytesIO(file_bytes), f.filename)

        if not students:
            flash("No student marks data found in file.", "error")
            return _redirect_admin_back("reports", report_dept=department, report_year=year_level)

        subjects = [s["name"] for s in test_info.subjects]
        if not subjects:
            flash("Upload blocked: no subject columns detected.", "warning")
            return _redirect_admin_back("reports", report_dept=department, report_year=year_level)

        student_data = [s.to_dict() for s in students]
        test_name = _normalize_allowed_test_name(test_name_input or test_info.test_name)
        if not test_name:
            flash("Test name must be one of: IAT 1, IAT 2, MODEL EXAM.", "error")
            return _redirect_admin_back("reports", report_dept=department, report_year=year_level)

        if not batch_name:
            batch_name = f"{datetime.now().year}-{str(datetime.now().year + 1)[-2:]}"

        existing = db.find_existing_department_year_test(
            department=department,
            year_level=year_level,
            semester=semester,
            test_name=test_name,
            batch_name=batch_name,
        )

        replace_test_id = None
        if existing:
            if (existing.get("file_hash") or "") == file_hash:
                flash("Duplicate file detected for this department/year/test. Upload blocked.", "warning")
                return _redirect_admin_back("reports", report_dept=department, report_year=year_level)
            if upload_mode == "replace":
                replace_test_id = int(existing.get("test_id"))

        ok, msg = db.save_test_marks(
            test_name=test_name,
            semester=semester,
            counselor_email=session["user_email"],
            students=student_data,
            subjects=subjects,
            batch_name=batch_name,
            department=department,
            section=section,
            file_hash=file_hash,
            replace_test_id=replace_test_id,
            sync_students=False,
            year_level=year_level,
            enforce_assigned_match=False,
            uploaded_by=session["user_email"],
        )
        if ok:
            flash(f"Marksheet uploaded for {department} Year {year_level} ({len(student_data)} students).", "success")
        else:
            flash(f"Upload failed: {msg}", "error")
    except Exception as e:
        flash(f"Upload failed: {e}", "error")

    return _redirect_admin_back("reports", report_dept=department, report_year=year_level)


@app.route("/api/tests/<int:test_id>/counselor-update", methods=["POST"])
@login_required
def api_counselor_update_test(test_id):
    email = session.get("user_email")
    role = session.get("role", "counselor")
    if _is_counselor_department_blocked(email, role):
        flash("Your department is blocked. Editing is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database"))
    if not _can_access_test_for_user(test_id, email, role):
        flash("Access denied for this test.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Editing is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database"))

    try:
        meta = db.get_test_metadata(test_id) or {}
        test_name_raw = (request.form.get("test_name") or "").strip() or (meta.get("test_name") or "")
        test_name = _normalize_allowed_test_name(test_name_raw)
        if not test_name:
            flash("Test name must be one of: IAT 1, IAT 2, MODEL EXAM.", "error")
            return redirect(url_for("counselor_page", tab="test-database"))
        semester = (request.form.get("semester") or "").strip() or (meta.get("semester") or "")
        batch_name = (request.form.get("batch_name") or "").strip() or (meta.get("batch_name") or "")
        section = (request.form.get("section") or "").strip() or (meta.get("section") or "")

        db.update_test_metadata_fields(
            test_id,
            test_name=test_name,
            semester=semester,
            batch_name=batch_name,
            section=section,
        )
        flash("Test details updated.", "success")
    except Exception as e:
        flash(f"Could not update test: {e}", "error")

    next_url = (request.form.get("next") or "").strip()
    if next_url:
        parsed = urlparse(next_url)
        if not parsed.scheme and not parsed.netloc and next_url.startswith("/"):
            return redirect(next_url)

    return redirect(url_for("counselor_page", tab="test-database"))


@app.route("/api/tests/<int:test_id>/marks")
@login_required
def api_get_test_marks(test_id):
    """Get test marks grouped by student for display."""
    try:
        user_email = session.get("user_email")
        role = session.get("role")
        if not _can_access_test_for_user(test_id, user_email, role):
            return jsonify({"success": False, "error": "Access denied for this test."}), 403
        if role == "counselor" and _is_test_blocked(test_id):
            return jsonify({"success": False, "error": "This test is blocked by administration."}), 403

        if role == "counselor":
            data = db.get_test_marks_grouped_for_counselor(test_id, user_email)
        else:
            data = db.get_test_marks_grouped(test_id)
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/counselor/tests/<int:test_id>/marks/update", methods=["POST"])
@login_required
def api_counselor_update_marks(test_id):
    email = session.get("user_email")
    role = session.get("role")
    if role != "counselor":
        return jsonify({"success": False, "error": "Counselor access required."}), 403
    if not _can_access_test_for_user(test_id, email, role):
        return jsonify({"success": False, "error": "Access denied."}), 403
    if _is_test_blocked(test_id):
        return jsonify({"success": False, "error": "This test is blocked by administration."}), 403
    if _is_counselor_department_blocked(email, role):
        return jsonify({"success": False, "error": "Department is blocked. Editing disabled."}), 403

    try:
        payload = request.get_json(force=True, silent=False) or {}
        reg_no = str(payload.get("reg_no") or "").strip()
        marks = payload.get("marks") or {}
        if not reg_no or not isinstance(marks, dict):
            return jsonify({"success": False, "error": "Invalid payload."}), 400

        for subject_name, value in marks.items():
            if not str(subject_name or "").strip():
                continue
            db.upsert_counselor_mark_override(email, test_id, reg_no, str(subject_name), str(value or ""))

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/tests/cleanup-duplicates", methods=["POST"])
@login_required
@admin_required
def api_tests_cleanup_duplicates():
    system_only = _ensure_system_admin("reports")
    if system_only:
        return system_only

    try:
        deleted_count = db.cleanup_duplicate_tests()
        if deleted_count > 0:
            flash(f"Removed {deleted_count} duplicate test(s).", "success")
        else:
            flash("No duplicate tests found.", "info")
    except Exception as e:
        flash(f"Failed to cleanup duplicates: {e}", "error")
    return _redirect_admin_back("reports")


@app.route("/api/marks/upload", methods=["POST"])
@login_required
def api_upload_marks():
    flash("Marksheet upload is now admin-only. Use Departments tab in Admin panel.", "warning")
    if _is_admin_portal_user(session.get("role")):
        return _redirect_admin_back("reports")
    return redirect(url_for("counselor_page", tab="test-database"))

    email = session["user_email"]
    user = db.get_user(email) or {}
    f = request.files.get("marks_file")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return redirect(url_for("counselor_page"))

    try:
        file_bytes = f.read()
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        upload_mode = request.form.get("upload_mode", "new")
        replace_test_id = request.form.get("replace_test_id", type=int)

        from core.intelligent_parser import IntelligentParser
        parser = IntelligentParser()
        test_info, students = parser.parse_file(io.BytesIO(file_bytes), f.filename)

        if not students:
            flash("No student marks data found in file.", "error")
            return redirect(url_for("counselor_page"))

        subjects = [s["name"] for s in test_info.subjects]
        student_data = [s.to_dict() for s in students]

        if not subjects:
            flash("UPLOAD BLOCKED: HEADER INTEGRITY FAILED - no subject columns detected. FILE NOT UPLOADED.", "warning")
            return redirect(url_for("counselor_page", tab="test-database"))

        def _norm_dept(value):
            return re.sub(r"[^A-Za-z0-9]", "", str(value or "").upper())

        def _norm_reg(value):
            reg = str(value or "").strip().replace(" ", "")
            if reg.endswith(".0"):
                reg = reg[:-2]
            return reg.upper()

        def _infer_department_from_regnos(rows):
            """Best-effort department inference using configured register patterns."""
            score = {}
            for st in rows:
                reg = _norm_reg(st.get("reg_no"))
                if not reg:
                    continue
                digits = "".join(ch for ch in reg if ch.isdigit())
                if not digits:
                    continue
                for dept_code, pattern in DEPT_REG_PATTERNS.items():
                    if str(pattern) and str(pattern) in digits:
                        score[dept_code] = score.get(dept_code, 0) + 1

            if not score:
                return ""
            winner = max(score, key=score.get)
            return winner

        counselor_department = (user.get("department") or "").strip()
        parsed_department = (test_info.department or "").strip()
        if not counselor_department:
            flash("UPLOAD BLOCKED: counselor department is not configured. Contact admin. FILE NOT UPLOADED.", "warning")
            return redirect(url_for("counselor_page", tab="test-database"))

        if not parsed_department:
            parsed_department = _infer_department_from_regnos(student_data)

        if not parsed_department:
            # Header formats vary a lot across departments; fallback keeps uploads usable
            # while counselor ownership checks below still prevent cross-department data injection.
            parsed_department = counselor_department
            flash("Header department not detected. Used your counselor department for validation.", "info")

        if _norm_dept(parsed_department) != _norm_dept(counselor_department):
            flash(
                f"UPLOAD BLOCKED: NO MATCH - marksheet department '{parsed_department}' does not match your department '{counselor_department}'. FILE NOT UPLOADED.",
                "warning",
            )
            return redirect(url_for("counselor_page", tab="test-database"))

        # Enforce counselor scope: only admin-assigned students whose Reg No + Name match.
        def _norm_name(value):
            name = str(value or "").strip().lower()
            name = re.sub(r"\s+", " ", name)
            return name

        assigned_students = db.get_students(email)
        assigned_map = {
            _norm_reg(s.get("reg_no")): _norm_name(s.get("student_name"))
            for s in assigned_students
            if _norm_reg(s.get("reg_no"))
        }

        matched_students = []
        mismatch_examples = []
        mismatch_count = 0
        for st in student_data:
            reg = _norm_reg(st.get("reg_no"))
            name = _norm_name(st.get("name"))
            assigned_name = assigned_map.get(reg)
            if not reg or not assigned_name:
                mismatch_count += 1
                if len(mismatch_examples) < 5:
                    mismatch_examples.append(f"{st.get('reg_no', '')} ({st.get('name', '')})")
                continue
            if assigned_name != name:
                mismatch_count += 1
                if len(mismatch_examples) < 5:
                    mismatch_examples.append(f"{st.get('reg_no', '')} ({st.get('name', '')})")
                continue
            st["reg_no"] = reg
            matched_students.append(st)

        if not matched_students:
            sample = ", ".join(mismatch_examples) if mismatch_examples else "No valid rows"
            flash(
                f"UPLOAD BLOCKED: NO MATCH - no assigned students matched by Reg No + Name. Examples: {sample}. FILE NOT UPLOADED.",
                "warning",
            )
            return redirect(url_for("counselor_page", tab="test-database"))

        total_rows = len(student_data)
        if mismatch_count > 0:
            sample = ", ".join(mismatch_examples) if mismatch_examples else "Check uploaded rows"
            flash(
                f"PARTIAL MATCH: scanned {total_rows} rows, matched {len(matched_students)}, skipped {mismatch_count} non-assigned/mismatched rows. Examples: {sample}.",
                "warning",
            )

        student_data = matched_students

        raw_test_name = (request.form.get("test_name") or test_info.test_name or "IAT 1").strip()
        test_name = _normalize_allowed_test_name(raw_test_name) or "IAT 1"

        semester = (request.form.get("semester") or str(test_info.semester or "1")).strip()
        batch_name = (request.form.get("batch_name") or test_info.batch_name or "").strip()
        section = (request.form.get("section") or test_info.section or "").strip()
        department = counselor_department

        existing_test = db.find_existing_department_test(department, semester, test_name, batch_name=batch_name)
        if existing_test and upload_mode != "replace":
            if (existing_test.get("file_hash") or "") == file_hash:
                flash("UPLOAD BLOCKED: DUPLICATE FILE for this department/test. FILE NOT UPLOADED.", "warning")
                return redirect(url_for("counselor_page", tab="test-database"))
            replace_test_id = int(existing_test.get("test_id"))

        if upload_mode == "replace" and not replace_test_id and existing_test:
            replace_test_id = int(existing_test.get("test_id"))

        ok, msg = db.save_test_marks(
            test_name,
            semester,
            email,
            student_data,
            subjects,
            batch_name=batch_name,
            department=department,
            section=section,
            file_hash=file_hash,
            replace_test_id=replace_test_id,
            sync_students=False,
        )
        if ok:
            # Automatic duplicate cleanup after upload.
            db.cleanup_duplicate_tests()
            latest_test_id = db.get_latest_test_id_for_counselor(email)
            verb = "updated" if replace_test_id else "uploaded"
            msg_text = f"Marks {verb} — scanned {total_rows} rows, uploaded {len(student_data)} assigned students, {len(subjects)} subjects."
            flash(msg_text, "success")
            return redirect(url_for("counselor_test_send_page", test_id=latest_test_id or replace_test_id))
        else:
            if "no match" in str(msg).lower():
                flash(msg, "warning")
            else:
                flash(f"Error: {msg}", "error")
    except Exception as e:
        flash(f"Parse error: {e}", "error")
    return redirect(url_for("counselor_page", tab="test-database"))


def _can_access_test_for_user(test_id: int, user_email: str, role: str) -> bool:
    if _is_system_admin(role) or _is_principal(role):
        return True
    if _is_hod(role) or _is_deo(role):
        meta = db.get_test_metadata(test_id) or {}
        return _can_manage_department_year(user_email, role, meta.get("department"), meta.get("year_level") or 1)
    tests = db.get_visible_tests_for_counselor(user_email)
    allowed_ids = {int(t.get("id")) for t in tests if t.get("id") is not None}
    return test_id in allowed_ids


def _is_test_blocked(test_id: int) -> bool:
    meta = db.get_test_metadata(test_id) or {}
    return bool(int(meta.get("is_blocked") or 0))


@app.route("/counselor/tests/<int:test_id>/view")
@login_required
def counselor_test_view_page(test_id):
    email = session.get("user_email")
    role = session.get("role", "counselor")
    if not _can_access_test_for_user(test_id, email, role):
        flash("Access denied for this test.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Viewing is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_counselor_department_blocked(email, role):
        flash("Your department is blocked. Contact system admin.", "warning")
        return redirect(url_for("counselor_page", tab="recent-tests"))

    test_meta = db.get_test_metadata(test_id) or {}
    grouped = db.get_test_marks_grouped_for_counselor(test_id, email) if role == "counselor" else db.get_test_marks_grouped(test_id)
    test_department = str(test_meta.get("department") or "").strip()
    if role == "counselor" and not test_department:
        user = db.get_user(email) or {}
        test_department = str(user.get("department") or "").strip()

    if not test_department:
        for student in grouped.get("students", []):
            candidate = str(student.get("department") or "").strip()
            if candidate:
                test_department = candidate
                break

    if test_department:
        test_meta["department"] = test_department

    if role == "counselor":
        def _norm_reg(value):
            reg = str(value or "").strip().replace(" ", "")
            if reg.endswith(".0"):
                reg = reg[:-2]
            return reg.upper()

        allowed_reg_nos = {_norm_reg(s.get("reg_no", "")) for s in db.get_students(email)}
        grouped["students"] = [
            s for s in grouped.get("students", [])
            if _norm_reg(s.get("reg_no", "")) in allowed_reg_nos
        ]

    if test_department:
        for student in grouped.get("students", []):
            if not str(student.get("department") or "").strip():
                student["department"] = test_department

    return render_template(
        "counselor_test_view.html",
        test_id=test_id,
        test_meta=test_meta,
        subjects=grouped.get("subjects", []),
        students=grouped.get("students", []),
    )


@app.route("/counselor/tests/<int:test_id>/send")
@login_required
def counselor_test_send_page(test_id):
    email = session.get("user_email")
    role = session.get("role", "counselor")
    if not _can_access_test_for_user(test_id, email, role):
        flash("Access denied for this test.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Sending is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_counselor_department_blocked(email, role):
        flash("Your department is blocked. Sending is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="recent-tests"))

    user = db.get_user(email)
    students = db.get_students(email)
    def _norm_reg(value):
        reg = str(value or "").strip().replace(" ", "")
        if reg.endswith(".0"):
            reg = reg[:-2]
        return reg.upper()

    by_reg = {_norm_reg(s.get("reg_no")): s for s in students}
    grouped = db.get_test_marks_grouped_for_counselor(test_id, email) if role == "counselor" else db.get_test_marks_grouped(test_id)
    test_meta = db.get_test_metadata(test_id) or {}
    sent_reg_nos = {_norm_reg(r) for r in db.get_sent_reg_nos_for_test(email, test_id)}

    rows = []
    for sm in grouped.get("students", []):
        reg_no = sm.get("reg_no")
        norm_reg = _norm_reg(reg_no)
        if role == "counselor" and norm_reg not in by_reg:
            continue
        stu = by_reg.get(norm_reg, {})
        rows.append({
            "reg_no": norm_reg,
            "student_name": stu.get("student_name") or sm.get("name") or reg_no,
            "parent_phone": stu.get("parent_phone", ""),
            "department": stu.get("department") or test_meta.get("department") or user.get("department") or "",
            "marks": sm.get("marks", {}),
            "status": "Generated" if norm_reg in sent_reg_nos else "Pending",
        })

    return render_template(
        "counselor_send_results.html",
        test_id=test_id,
        test_meta=test_meta,
        rows=rows,
        country_code=COUNTRY_CODE,
    )


@app.route("/api/reports/send-single", methods=["POST"])
@login_required
def api_send_single_report():
    email = session.get("user_email")
    role = session.get("role", "counselor")
    test_id = request.form.get("test_id", type=int)
    reg_no = request.form.get("reg_no", "").strip()
    action = request.form.get("action", "cancel")
    is_ajax = request.form.get("ajax") == "1"
    ordered_fields_raw = request.form.get("ordered_fields", "").strip()

    ordered_fields = None
    if ordered_fields_raw:
        try:
            parsed = json.loads(ordered_fields_raw)
            if isinstance(parsed, list):
                ordered_fields = parsed
        except Exception:
            ordered_fields = None

    if not test_id or not reg_no:
        flash("Test and student are required.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))

    if not _can_access_test_for_user(test_id, email, role):
        flash("Access denied for this test.", "error")
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Sending is disabled.", "warning")
        if is_ajax:
            return jsonify({"success": False, "error": "Test is blocked."}), 403
        return redirect(url_for("counselor_page", tab="test-database"))
    if _is_counselor_department_blocked(email, role):
        flash("Your department is blocked. Sending is disabled.", "warning")
        if is_ajax:
            return jsonify({"success": False, "error": "Department is blocked."}), 403
        return redirect(url_for("counselor_page", tab="recent-tests"))

    if action != "send":
        flash("Message status kept as Pending.", "info")
        if is_ajax:
            return jsonify({"success": True, "status": "pending"})
        return redirect(url_for("counselor_test_send_page", test_id=test_id))

    user = db.get_user(email)
    test_meta = db.get_test_metadata(test_id) or {}
    students = db.get_students(email)
    def _norm_reg(value):
        reg = str(value or "").strip().replace(" ", "")
        if reg.endswith(".0"):
            reg = reg[:-2]
        return reg.upper()

    normalized_reg = _norm_reg(reg_no)
    stu = next((s for s in students if _norm_reg(s.get("reg_no")) == normalized_reg), None)
    if not stu:
        flash("Student not found under your account.", "error")
        return redirect(url_for("counselor_test_send_page", test_id=test_id))

    if role == "counselor":
        marks = db.get_student_marks_for_reg_for_counselor(test_id, normalized_reg, email)
    else:
        marks = db.get_student_marks_for_reg(test_id, normalized_reg)
    if not marks:
        flash("No marks found for selected student.", "error")
        return redirect(url_for("counselor_test_send_page", test_id=test_id))

    from utils.whatsapp_helper import get_whatsapp_link
    from utils.template_engine import TemplateEngine

    template = request.form.get("message_template", "").strip() or (
        "Dear Parent , The Following is the {test_name} Marks Secured in each Course by your son/daughter\n\n"
        "REGISTER NUMBER :  {reg_no}\n"
        "NAME : {student_name}\n\n"
        "{subjects_table}\n\n"
        "Regards\n"
        "PRINCIPAL\n"
        "RMKCET"
    )

    marks_table = _build_parent_subjects_table(marks, ordered_fields=ordered_fields)
    effective_test_name = request.form.get("test_name") or (test_meta.get("test_name") or "Unit Test")
    if request.form.get("message_template", "").strip():
        msg = TemplateEngine.fill_template(
            template,
            app_name=APP_NAME,
            reg_no=normalized_reg,
            student_name=stu.get("student_name", normalized_reg),
            department=request.form.get("department") or (test_meta.get("department") or stu.get("department", "")),
            test_name=effective_test_name,
            semester=request.form.get("semester") or (test_meta.get("semester") or "-"),
            batch_name=request.form.get("batch_name") or (test_meta.get("batch_name") or "-"),
            section=request.form.get("section") or (test_meta.get("section") or "-"),
            subjects_table=marks_table,
            counselor_name=user.get("name", "Counselor"),
        )
    else:
        msg = _build_parent_message(
            effective_test_name,
            normalized_reg,
            stu.get("student_name", normalized_reg),
            marks,
        )

    def _clean_phone(value):
        digits = "".join(ch for ch in str(value or "") if ch.isdigit())
        return digits[-10:] if len(digits) >= 10 else ""

    def _clean_reg(value):
        reg = str(value or "").strip().replace(" ", "")
        if reg.endswith(".0"):
            reg = reg[:-2]
        return reg

    phone = _clean_phone(stu.get("parent_phone", ""))
    if not phone:
        # Fallback for older uploads where phone may have been parsed into email-like field.
        fallback_phone = _clean_phone(stu.get("parent_email", ""))
        if fallback_phone:
            phone = fallback_phone
            try:
                db.update_student(email, reg_no, parent_phone=phone)
            except Exception:
                pass

    if not phone:
        # Secondary fallback: recover from any duplicate/legacy student row with equivalent reg number.
        target_reg = _clean_reg(reg_no)
        for other in students:
            if _clean_reg(other.get("reg_no")) != target_reg:
                continue
            alt_phone = _clean_phone(other.get("parent_phone")) or _clean_phone(other.get("parent_email"))
            if alt_phone:
                phone = alt_phone
                try:
                    db.update_student(email, reg_no, parent_phone=phone)
                except Exception:
                    pass
                break

    if not phone:
        flash(f"Parent phone number missing for {reg_no}.", "error")
        if is_ajax:
            return jsonify({"success": False, "error": f"Parent phone number missing for {normalized_reg}."}), 400
        return redirect(url_for("counselor_test_send_page", test_id=test_id))

    wa = get_whatsapp_link(phone, msg)
    db.log_message(email, normalized_reg, stu.get("student_name", ""), msg, "message", wa, test_id=test_id)

    if is_ajax:
        return jsonify({"success": True, "status": "generated", "wa_link": wa})

    # Open WhatsApp compose URL so counselor can send immediately.
    return redirect(wa)


@app.route("/api/reports/generate", methods=["POST"])
@login_required
def api_generate_reports():
    email = session["user_email"]
    role = session.get("role", "counselor")
    test_id = request.form.get("test_id")
    reg_nos = request.form.getlist("reg_nos")
    fmt = request.form.get("format", "message")

    # Editable parsed fields
    edited_test_name = request.form.get("edited_test_name", "").strip()
    edited_semester = request.form.get("edited_semester", "").strip()
    edited_department = request.form.get("edited_department", "").strip()
    edited_batch = request.form.get("edited_batch", "").strip()
    custom_message_body = request.form.get("custom_message_body", "").strip()

    if not test_id or not reg_nos:
        flash("Select a test and at least one student.", "error")
        return redirect(url_for("counselor_page", tab="test-database", test_id=test_id or ""))

    test_id = int(test_id)

    if not _can_access_test_for_user(test_id, email, role):
        flash("Access denied for this test.", "error")
        return redirect(url_for("counselor_page", tab="test-database", test_id=test_id))

    if _is_test_blocked(test_id):
        flash("This test is blocked by administration. Sending is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="test-database", test_id=test_id))

    if _is_counselor_department_blocked(email, role):
        flash("Your department is blocked. Sending is disabled.", "warning")
        return redirect(url_for("counselor_page", tab="recent-tests"))

    user = db.get_user(email)
    test_meta = db.get_test_metadata(test_id)

    # Keep metadata editable before send
    if edited_test_name or edited_semester or edited_department or edited_batch:
        db.update_test_metadata_fields(
            test_id,
            test_name=edited_test_name or (test_meta or {}).get("test_name") or "",
            semester=edited_semester or (test_meta or {}).get("semester") or "",
            department=edited_department or (test_meta or {}).get("department") or "",
            batch_name=edited_batch or (test_meta or {}).get("batch_name") or "",
        )
        test_meta = db.get_test_metadata(test_id)

    students = db.get_students(email)
    lookup = {s["reg_no"]: s for s in students}

    from utils.whatsapp_helper import get_whatsapp_link
    from utils.template_engine import TemplateEngine

    already_sent = db.get_sent_reg_nos_for_test(email, test_id)
    reports = []
    for rn in reg_nos:
        if role == "counselor" and rn not in lookup:
            continue
        if rn in already_sent:
            continue

        if role == "counselor":
            marks = db.get_student_marks_for_reg_for_counselor(test_id, rn, email)
        else:
            marks = db.get_student_marks_for_reg(test_id, rn)
        stu = lookup.get(rn, {})
        if not marks:
            continue

        marks_table = _build_parent_subjects_table(marks)
        test_name = (test_meta or {}).get("test_name") or "Unit Test"
        semester = (test_meta or {}).get("semester") or "-"
        department = (test_meta or {}).get("department") or stu.get("department", "-")
        batch_name = (test_meta or {}).get("batch_name") or "-"

        if custom_message_body:
            msg = TemplateEngine.fill_template(
                custom_message_body,
                app_name=APP_NAME,
                reg_no=rn,
                student_name=stu.get("student_name", rn),
                department=department,
                test_name=test_name,
                semester=semester,
                batch_name=batch_name,
                subjects_table=marks_table,
                counselor_name=user["name"],
            )
        else:
            msg = _build_parent_message(test_name, rn, stu.get('student_name', rn), marks)

        phone = stu.get("parent_phone", "")
        wa = get_whatsapp_link(phone, msg) if phone else ""
        db.log_message(email, rn, stu.get("student_name", ""), msg, fmt, wa, test_id=test_id)

        reports.append({
            "reg_no": rn,
            "name": stu.get("student_name", rn),
            "marks": marks,
            "message": msg,
            "whatsapp_link": wa,
            "phone": phone,
        })

    session["reports"] = reports
    session["report_test_id"] = test_id
    flash(f"Reports generated for {len(reports)} pending students.", "success")
    return redirect(url_for("counselor_page", tab="test-database", test_id=test_id))


@app.route("/api/reports/pdf/<reg_no>")
@login_required
def api_student_pdf(reg_no):
    email = session["user_email"]
    role = session.get("role", "counselor")
    test_id = request.args.get("test_id")
    if not test_id:
        return "Missing test_id", 400

    test_id = int(test_id)
    if not _can_access_test_for_user(test_id, email, role):
        return "Access denied for this test", 403

    if _is_test_blocked(test_id):
        return "This test is blocked", 403

    if _is_counselor_department_blocked(email, role):
        return "Department is blocked", 403

    if role == "counselor":
        marks = db.get_student_marks_for_reg_for_counselor(test_id, reg_no, email)
    else:
        marks = db.get_student_marks_for_reg(test_id, reg_no)
    if not marks:
        return "No marks found", 404

    stu = next((s for s in db.get_students(email) if s["reg_no"] == reg_no), None)
    if role == "counselor" and not stu:
        return "Student not found under your account", 403
    user = db.get_user(email)
    meta = db.get_test_metadata(test_id)

    from utils.pdf_generator import generate_student_pdf
    pdf_bytes = generate_student_pdf(
        student_name=stu["student_name"] if stu else reg_no,
        reg_no=reg_no,
        department=(stu or {}).get("department", ""),
        subjects_marks=marks,
        counselor_name=user["name"],
        test_name=(meta or {}).get("test_name", "Unit Test"),
    )
    return send_file(
        io.BytesIO(pdf_bytes), mimetype="application/pdf",
        as_attachment=True, download_name=f"{reg_no}_report.pdf",
    )


# ---------- Format Settings -------------------------------------------------

@app.route("/api/settings/format", methods=["POST"])
@login_required
@admin_required
def api_update_format_settings():
    system_only = _ensure_system_admin("config")
    if system_only:
        return system_only

    default = request.form.get("default_format", "message")
    allowed = request.form.getlist("allowed_formats")
    bulk = request.form.get("bulk_format", "same_as_individual")
    db.update_format_settings(default, allowed, bulk, session["user_email"])
    flash("Format settings updated.", "success")
    return _redirect_admin_back("config")


# ---------- Static data assets ----------------------------------------------

@app.route("/data/<path:filename>")
@login_required
@admin_required
def serve_data(filename):
    system_only = _ensure_system_admin("reports")
    if system_only:
        return system_only

    safe_name = os.path.normpath(filename).lstrip("\\/")
    candidate = os.path.abspath(os.path.join(DATA_DIR, safe_name))
    data_root = os.path.abspath(DATA_DIR)

    if not (candidate == data_root or candidate.startswith(data_root + os.sep)):
        abort(404)
    if not os.path.isfile(candidate):
        abort(404)

    return send_file(candidate)


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 5000
    print(f"\n  ✅  RMKCET Parent Connect running at: http://localhost:{port}\n")
    app.run(debug=False, host="0.0.0.0", port=port, use_reloader=False)
